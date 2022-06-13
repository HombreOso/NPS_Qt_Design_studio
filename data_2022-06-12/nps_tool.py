import gc
import re
import tkinter as tk
from tkinter import *
from tkinter.filedialog import askdirectory
import numpy as np
import os
import xlsxwriter as xlsx
import openpyxl as opxl
from natsort import natsorted, ns
import pydicom
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('TkAgg')
from PIL import Image, ImageTk
import itertools
import cv2
import time
import functools as fut
import pandas as pd
import docx
from scipy import optimize as opt
import multiprocessing as mp
import json


class StartClass:

    """
    Start of the program.
    
    Create auxiliary folders;
    Create file list and file dict;
    Create window for specifying metadata to be extracted;
    Define static method create_base_array;
    Create PNG-images of dicoms to be shown in GUI.

    Attributes
        ----------
        acceptButtonIsAlreadyUsed : boolean
            Auxiliary boolean
            True: auto-fill of 'fixed'-tool dialog window with last
            set size of rectangular ROI;
            False: use values defined in GUI class' init (by default: 128x128)

        program_start : boolean
            Auxiliary boolean
            True: when creating an object of class CreateForm, the form for
            file exclusion is created;
            False: other forms are created depending on the value of tool_index
            attribute of an object of class GUI

        aux_folder_names : list of strings
            Names of auxiliary folders to be created at the start of the program:
                '01.2d_NPS_images': folder with 2d_NPS images (.png) of the last calculated ROIs;
                '02.One_D_NPS': matplotlib charts of 1d-NPS profiles saved as .png;
                '03.PNG_images': Dicoms that are transformed in .png-format, to be shown in GUI.

        whole_paths : list of strings
            Absolute paths to created auxiliary folders.

        file_exclusion_json : string
            Name of JSON containing last settings for file exclusion.

        im_height_dict : dict
            Keys : absolute paths to dcm-images;
            Values : the heights (in pixels) of the respective images.

        im_width_dict : dict
            Keys : absolute paths to dcm-images;
            Values : the widths (in pixels) of the respective images.

        metadata_tags_list : list of lists of hexstrings
            Each inner list contains even number of elements defining either tag (two elements) or
            tag and subtags (four and more elements). Each element is string of a hexadecimal number.
            For example: [['0x0008', '0x1030'], ['0x0040', '0x0275', '0x0032', '0x1060']] defines two tags:
            the second one is the subtag of tag ['0x0040']['0x0275']

        metadata_subdict : dict
            Dictionary based on metadata tag list, retrieving specified metadata from dicoms
            Keys : tag names;
            Values : values of tags.

        metadata_dict : dict of dicts
            Keys : absolute paths to dcm_images;
            Values : respective metadata_subdict.

        folder_with_images : string
            Absolute path to folder with all dicoms to be analyzed.

        filelist : list of strings
            Absolute paths to all images to be analyzed.

        filedict : dict of dicts
            Dict for sorting the attribute filelist.
            Outer dict:
                Keys : absolute paths to studies' folders;
                Values : inner dicts.
            Inner dict:
                Keys : absolute paths to series' folders;
                Values : list of files that are in respective series' folder.

        new_files : dict of dicts
            Dict containing filelist and filedict attributes.

        num_files_to_exclude_start : int
            Number of files to be excluded from the beginning of each series' folder.

        num_files_to_exclude_end : int
            Number of files to be excluded from the end of each series' folder.

        all_images : list of strings
            Absolute paths to all png-images of dicoms.

        Methods
        -------
        select_folder(self, title)
            Create dialog window for selecting folder with all images
            to be analyzed. Used to build attribute folder_with_images.

        create_dataset_dictionary(list_of_indices, dataset_dicom)
            Retrieve metadata from current dicom-file. Used to build attribute metadata_subdict.

        create_filelist(self, pathtoFiles, suffix_array)
            Search for files with specified extensions and build list of them.

        create_filedict(self, pathtoFiles, suffix_array)
            Search for files with specified extensions and build
            sorting dict of files to be analyzed.

        exclude_files(self, file_dict, file_list,
                      num_files_to_exclude_start,
                      num_files_to_exclude_end)
            Exclude specified number of files from each series folder
            and perform changes in filedict and filelist attributes
            of class StartClass.

        create_base_array(self, image_file)
            Read current dicom file and retrieve pixel array.
            Retrieve part of meta data and update attribute metadata_dict.

        rgb2gray(self, rgb)
            If we handle with RGB-image, convert it to grayscale.

        @staticmethod
        norm_2d_array(arr_2d, value_min, value_max)
            Normalize 2d-array to values between value_min and value_max.

        create_image_from_2d_array(self, arr_2d, filename)
            Normalize passed 2d_array between values 0 and 255 and create
            image saved as filename.

        create_png_image(self, key)
            Create .png-image from current dicom-file.

        @staticmethod
        create_aux_folder(folder_name)
            Create auxiliary folder in the directory with executable py-file.

        @staticmethod
        recognize_hex_numbers_in_string(settings_string)
            Recognize hexadecimal numbers in given string.



       """

    def __init__(self, suffixes, list_of_indices_raw):
        """ 
        :param suffixes: list of strings
            Suffixes of files to be found in selected directory
            (DICOMs or image files)
        :param list_of_indices_raw: list of strings
            Default list of tags of metadata to be extracted from DICOMs
            Specified in init_dict


        """

        print('Constructor of the class StartClass is being executed')
        # extensions of files to be found
        self.suffixes = suffixes
        # default list of tags for metadata
        self.list_of_indices_raw = list_of_indices_raw

        # auxiliary booleans
        # global acceptButtonIsAlreadyUsed
        # acceptButtonIsAlreadyUsed = False
        self.acceptButtonIsAlreadyUsed = False
        self.program_start = True

        # create auxiliary folders
        self.aux_folders_names = ['01.2d_NPS_images',
                                  '02.One_D_NPS',
                                  '03.PNG_images']
        self.whole_paths = list(map(StartClass.create_aux_folder, self.aux_folders_names))

        # name for json-file containing previous settings
        self.file_exclusion_json = 'file_exclusion_settings.txt'
        # dictionary for measurements of each image
        self.im_height_dict = {}
        self.im_width_dict = {}
        # initialize dictionary containing meta data
        # keys: paths to dicoms
        # values: part dictionaries with meta data
        self.metadata_dict = {}
        # collect all images in png-format
        self.all_images = []

        # create dialog window to specify part of meta data to be retrieved
        # creating dialog window for entering tag numbers
        dialog_window = tk.Tk()
        dialog_window.geometry('%dx%d+%d+%d' % (init_dict['main_window_width_md'],
                                                init_dict['main_window_height_md'],
                                                init_dict['left_upper_corner_x_md'],
                                                init_dict['left_upper_corner_y_md']
                                                ))
        # instantiate object of class CreateFormMetaData
        obj_create_form = CreateFormMetaData(title='Tag Numbers',
                                             value=init_dict['list_of_indices_raw'],
                                             master_main_window=dialog_window)
        dialog_window.mainloop()

        # recognize tag numbers
        self.metadata_tags_list = StartClass.recognize_hex_numbers_in_string(
            settings_string=obj_create_form.settings_string)

        # select folder with images
        self.folder_with_images = self.select_folder(
            title='Select folder with images')
        # create list of found images
        self.filelist = self.create_filelist(pathtoFiles=self.folder_with_images,
                                             suffix_array=self.suffixes)

        self.filedict = self.create_filedict(pathtoFiles=self.folder_with_images,
                                             suffix_array=self.suffixes)

        # create dialog window to exclude some files from folders
        fields = ['remove_begin', 'remove_end']
        values = [3, 3]
        if os.path.isfile(os.path.join(os.getcwd(), self.file_exclusion_json)):
            with open(self.file_exclusion_json, 'r') as file_to_read_info:
                read_dict = json.load(file_to_read_info)
            list_of_variables = [read_dict[key] for key in read_dict]
            values = list_of_variables

        main_window = tk.Tk()
        obj_exclude = CreateForm(master_main_window=main_window, object_gui=None,
                                 object_arrays=None, fields=fields,
                                 values=values, program_start=self.program_start)
        main_window.mainloop()
        self.num_files_to_exclude_start = obj_exclude.num_files_to_exclude_start
        self.num_files_to_exclude_end = obj_exclude.num_files_to_exclude_end
        self.program_start = False

        # exclude files from folders
        
        self.new_files = self.exclude_files(file_list=self.filelist, file_dict=self.filedict,
                                            num_files_to_exclude_start=self.num_files_to_exclude_start,
                                            num_files_to_exclude_end=self.num_files_to_exclude_end)
        self.filelist = self.new_files['file_list']
        self.filedict = self.new_files['file_dict']

        # self.create_image_arrays(filelist=self.filelist)
        # create png-images
        for num, key in enumerate(self.filelist):
            if num > 1:
                break
            self.all_images.append(self.create_png_image(key=key))

        print('Constructor of the class StartClass is done')

        pass

    @staticmethod
    def recognize_hex_numbers_in_string(settings_string):
        """
        Recognize hexadecimal numbers in given string.
        Used to build attribute metadata_tags_list.
        :param settings_string: string
            User input of tag numbers in dialog window.
        :return: list of hexadecimal strings
            (See attribute metadata_tags_list of class StartClass)
        """
        # initialize list for tag numbers
        list_of_indices = []
        # pattern for recognizing hex numbers using RegEx
        pattern = r'0x[\d\w]+'
        # iterate over lines in settings_string
        for string_line in settings_string.split('\n'):
            list_of_indices.append(re.findall(pattern=pattern,
                                              string=string_line))
        return list_of_indices

    @staticmethod
    def create_dataset_dictionary(list_of_indices, dataset_dicom):
        """
        Retrieve specified metadata from dicom-file.

        :param list_of_indices: list of hexstrings
            Tags' hexadecimal numbers. Attribute metadata_tags_list is a value
            of this parameter (see init of class StartClass).
        :param dataset_dicom: Dataset object
            Dataset object of current dicom-file
        :return: dataset_dict : dict
            (See attribute metadata_subdict in init of class StartClass).
        """

        # initialize dataset_dictionary
        dataset_dictionary = {}
        # iterate over indices in the passed list
        for prop_index in list_of_indices:
            if len(prop_index) < 2:
                continue
            first_index = prop_index[0]
            second_index = prop_index[1]
            # retrieve DataElement object
            try:
                d_element = dataset_dicom[first_index, second_index]
                # if sub indices are used
                if len(prop_index) > 2:
                    # iterate over indices in prop_index
                    for num_index, sub_index in enumerate(prop_index[2:]):
                        # iterate only over even indices of prop_index
                        if num_index % 2 != 0:
                            continue
                        d_element = d_element[0][sub_index,
                                                 prop_index[num_index + 3]]

                # get value of property
                value_of_property = d_element.value
                # get the name of property
                try:
                    name_of_property = d_element.name
                except IndexError:
                    name_of_property = ''
                    print('problem with: ', d_element)
                dataset_dictionary.update({name_of_property: value_of_property})
            except KeyError:
                dataset_dictionary.update({'undefined_tag': 'undefined'})
        return dataset_dictionary

    def select_folder(self, title):
        """
        Create dialog window for selecting folder with all images
        to be analyzed.
        :param title: titlle of dialog window;
        :return:
        Absolute path to folder with all images to be analyzed
        (See attribute folder_with_images of class StartClass)
        """
        root_2 = tk.Tk()
        selected_folder = askdirectory(master=root_2, title=title)
        root_2.destroy()
        root_2.mainloop()
        return selected_folder

    def create_filelist(self, pathtoFiles, suffix_array):
        """
        Search for files with specified extensions and build list of files to be analyzed.
        :param pathtoFiles: string
            Path to folder with all images to be analyzed
        :param suffix_array: list of strings
            Extensions of files to be searched for.
        :return: list of strings
            List of paths to all found images
        """

        print('create_filelist is being executed')
        # list of paths to all found files
        lstFiles = []
        # find out how many files found
        counter_files = 0
        sorted_file_names = []
        # list of all directories with found files
        all_dir_names = []
        # empty dict to sort files in directories
        directories_dict = {}
        # subdict = {}
        # search for files with certain extensions
        for dirName, subdirList, fileList in os.walk(pathtoFiles):
            for filename in fileList:
                # if any of extensions are present in filename
                if any(suffix.lower() in filename.lower() for suffix in suffix_array):
                    lstFiles.append(os.path.join(dirName, filename))
                    filepath = os.path.join(dirName, filename)
                    # base name of current directory with file
                    basedirname = os.path.basename(dirName)
                    classdirname = os.path.dirname(os.path.dirname(dirName))
                    try:  # update dict if the key-value-pair already exists
                        temp_dict = directories_dict[classdirname]
                    except KeyError:
                        sub_dict = {}
                        directories_dict.update({classdirname: sub_dict})
                        temp_dict = directories_dict[classdirname]
                    try:
                        temp_list = temp_dict[basedirname]
                        temp_list.append(filepath)
                        temp_dict.update({basedirname: temp_list})
                        directories_dict.update({classdirname: temp_dict})
                    except KeyError:
                        # or create key-value pair

                        # get subdict
                        temp_subdict = directories_dict[classdirname]
                        temp_subdict.update({basedirname: [filepath]})
                        directories_dict.update({classdirname: temp_subdict})

        # create array of only file names, sorted generally (based on whole path)
        for fnamepath in natsorted(lstFiles, alg=ns.IGNORECASE):
            sorted_file_names.append(os.path.basename(fnamepath))
        # create array of only folder names, sorted generally (based on whole path)
        for fdir in natsorted(lstFiles, alg=ns.IGNORECASE):
            all_dir_names.append(os.path.basename(os.path.dirname(fdir)))
        # print base names of found files
        for q in natsorted(lstFiles, alg=ns.IGNORECASE):
            print(os.path.basename(q))
        # print the number of found files
        print('%d files have been found' % len(lstFiles))
        print('create_filelist is done')
        return natsorted(lstFiles, alg=ns.IGNORECASE, key=lambda x: x.split('_')[-1])

    def create_filedict(self, pathtoFiles, suffix_array):

        """
        Search for files with specified extensions and build
        sorting dict of files to be analyzed.
        :param pathtoFiles: string
            Path to folder with all images to be analyzed
        :param suffix_array: list of strings
            Extensions of files to be searched for.
        :return: dict of dicts
            (See attribute filedict of class StartClass)
        """

        print('create_filedict is being executed')


        # list of paths to all found files
        lstFiles = []
        # find out how many files found
        counter_files = 0
        sorted_file_names = []
        # list of all directories with found files
        all_dir_names = []
        # empty dict to sort files in directories
        directories_dict = {}
        # subdict = {}
        # search for files with certain extensions
        for dirName, subdirList, fileList in os.walk(pathtoFiles):
            for filename in fileList:
                # # counter used to exclude files from the start of the list
                # counter_exclude_start = 1
                # # counter used to exclude file from the end of the list
                # counter_exclude_end = len(fileList)
                # if any of extensions are present in filename
                if any(suffix.lower() in filename.lower() for suffix in suffix_array):
                    lstFiles.append(os.path.join(dirName, filename))
                    filepath = os.path.join(dirName, filename)
                    # base name of current directory with file
                    basedirname = os.path.basename(dirName)
                    classdirname = os.path.dirname(dirName)
                    try:  # update dict if the key-value-pair already exists
                        temp_dict = directories_dict[classdirname]
                    except KeyError:
                        sub_dict = {}
                        directories_dict.update({classdirname: sub_dict})
                        temp_dict = directories_dict[classdirname]
                    try:
                        temp_list = temp_dict[basedirname]
                        temp_list.append(filepath)
                        temp_dict.update({basedirname: temp_list})
                        directories_dict.update({classdirname: temp_dict})
                    except KeyError:
                        # or create key-value-pair

                        # get subdict
                        temp_subdict = directories_dict[classdirname]
                        temp_subdict.update({basedirname: [filepath]})
                        directories_dict.update({classdirname: temp_subdict})

        # create array of only file names, sorted generally (based on whole path)
        for fnamepath in natsorted(lstFiles, alg=ns.IGNORECASE):
            sorted_file_names.append(os.path.basename(fnamepath))
        # create array of only folder names, sorted generally (based on whole path)
        for fdir in natsorted(lstFiles, alg=ns.IGNORECASE):
            all_dir_names.append(os.path.basename(os.path.dirname(fdir)))
        # print base names of found files
        # for q in natsorted(lstFiles, alg=ns.IGNORECASE):
        #     print(os.path.basename(q))
        # # print the number of found files
        # print('%d files have been found' % len(lstFiles))
        print('create_filedict is done')

        return directories_dict

    def exclude_files(self, file_dict, file_list,
                      num_files_to_exclude_start,
                      num_files_to_exclude_end):

        """
        Exclude specified number of files from each series folder
        and perform changes in filedict and filelist attributes
        of class StartClass.
        :param file_dict: dict of dict
            Attribute filedict before the exclusion of files.
        :param file_list: list of strings
            Attribute filelist before the exclusion of files.
        :param num_files_to_exclude_start: int
            (See attribute with the same name of class StartClass)
        :param num_files_to_exclude_end:
            (See attribute with the same name of class StartClass)
        :return: dict
            (See attribute new_files of class StartClass)
        """

        print('exclude_files is being executed')

        if num_files_to_exclude_end == 0 and num_files_to_exclude_start == 0:
            ret_dict = {'file_list': file_list,
                        'file_dict': file_dict}

            print('exclude_files is done')

            return ret_dict

        # empty list for files to exclude
        files_to_exclude = []

        # drop files in file_dict

        # iterate over folders
        for folder in natsorted(file_dict.keys(), key=lambda f: f.split('_')[-1]):
            temp_dict = file_dict[folder]
            # iterate over series
            for series in natsorted(temp_dict.keys(), key=lambda f: f.split('_')[-1]):
                # get list of files in series
                temp_list = temp_dict[series]
                # drop files
                try:
                    new_temp_list = temp_list[num_files_to_exclude_start: -num_files_to_exclude_end]
                    files_to_exclude += temp_list[:num_files_to_exclude_start]
                    files_to_exclude += temp_list[-num_files_to_exclude_end:]
                    temp_dict.update({series: new_temp_list})
                except IndexError:
                    print('There less files in the folder %s/%s, than attempted to exclude' % (folder, series))
            file_dict.update({folder: temp_dict})

        # get rid of repeated file names
        files_to_exclude = set(files_to_exclude)
        files_to_exclude = list(files_to_exclude)

        for file_to_remove in files_to_exclude:
            file_list.remove(file_to_remove)

        file_list = natsorted(file_list, alg=ns.IGNORECASE)

        ret_dict = {'file_list': file_list,
                    'file_dict': file_dict}

        print('exclude_files is done')

        return ret_dict

    def create_base_array(self, image_file):

        """
        Read current dicom file and retrieve pixel array.
        Retrieve part of meta data
        and update attribute metadata_dict
        :param image_file: string
            Absolute path to current image.
        :return: dict
            Key: 'base_array' : Value: pixel array of current image;
            Key: 'meatdata_subdict' : Value: dict of specified metadata;
            Key: 'whole_dcm' : Value: Dataset object of current dicom.
        """

        print('create_base_array is being executed')

        # if we handle with dicom-file
        if os.path.basename(image_file)[-4:] == '.dcm':
            try:
                # create data element object from dicom
                image_dcm = pydicom.dcmread(image_file, force=True)
            except:
                print('\n\n\nThere is a problem with the file: ')
                print(image_file)
            gc.collect()
            self.array = image_dcm.pixel_array
            self.metadata_subdict = StartClass.create_dataset_dictionary(
                list_of_indices=self.metadata_tags_list,
                dataset_dicom=image_dcm
            )
            metadata_subdict = self.metadata_subdict
            self.metadata_dict.update({image_file: self.metadata_subdict})

        # if we handle file with another file-extension
        else:
            # read image as list with PIL-library
            img = Image.open(image_file)
            # convert list into numpy-array
            self.array = np.array(img)
            # if we have colored image
            if len(self.array.shape) > 2:
                self.array = self.rgb2gray(self.array)
            self.metadata_subdict = {'undefined': 'undefined'}
            metadata_subdict = {'undefined': 'undefined'}

            # if the image is not a dicom, store 'undefined' in metadata_dict
            self.metadata_dict.update({image_file: {'undefined_tag': 'undefined'}})
            image_dcm = ''
        # image measurements
        self.px_height = self.array.shape[0]
        self.px_width = self.array.shape[1]
        # image file base name without extension
        self.basename = os.path.basename(image_file)[:-4]
        # image file base name with extension
        self.basename_w_ext = os.path.basename(image_file)

        ret_dict = {'base_array': self.array.astype(np.int16),
                    'metadata_subdict': metadata_subdict,
                    'whole_dcm': image_dcm}

        print('create_base_array is done')

        return ret_dict

    def rgb2gray(self, rgb):
        
        """
        If we handle with RGB-image, convert it to grayscale.
        :param rgb: ndarray
            ndarray (3d) of image to be transformed.
        :return: ndarray
            ndarray (2d) of the converted image
        """

        print('rgb2gray is being executed')

        # formula for converting of colored image to grayscale

        print('rgb2gray is done')

        return np.dot(rgb[..., :3], [0.299, 0.587, 0.114])

    @staticmethod
    def norm_2d_array(arr_2d, value_min, value_max):
        
        """
        Normalize 2d-array to values between value_min and value_max.
        :param arr_2d: ndarray (2d)
            Array to be normalized.
        :param value_min: int or float 
            Minimal value of normalized array.
        :param value_max: int or float
            Maximal value of normalized array.
        :return: ndarray (2d)
            Noirmalized array.
        """

        print('norm_2d_array is being executed')

        max_value = np.max(arr_2d)
        min_value = np.min(arr_2d)
        normalized_2d_array = value_min + (np.array(arr_2d) - min_value) / (max_value - min_value) * value_max

        print('norm_2d_array is done')

        return normalized_2d_array

    @staticmethod
    def create_image_from_2d_array(arr_2d, filename):

        """
        Normalize passed 2d_array between values 0 and 255 and create image saved as filename.

        :param arr_2d: ndarray (2d)
            Array the image to be created from.
        :param filename: string
            Absolute path to created image.
        :return: string
            filename
        """

        print('create_image_from_2d_array is being executed')

        image_array = np.array(StartClass.norm_2d_array(arr_2d=arr_2d,
                                                        value_min=0,
                                                        value_max=255))
        cv2.imwrite(filename=filename, img=image_array)

        print('create_image_from_2d_array is done')

        return filename

    def create_png_image(self, key):
        
        """
        Create .png-image from current dicom-file.
        :param key: string
            Absolute path to dicom file
        :return: string
            Absolute path to created image.
        """

        print('create_png_image is being executed')

        # pixel array for current image
        curr_array = self.create_base_array(key)['base_array']
        # name of image file without extension
        name_of_file_without_extension = os.path.basename(key)[:-4]
        # create png-image
        png_image_path = StartClass.create_image_from_2d_array(arr_2d=curr_array,
                                                               filename='03.PNG_images/%s.png' %
                                                               name_of_file_without_extension)
        # store path to png image in list
        # self.all_images.append(png_image_path)
        # return self.all_images

        print('create_png_image is done')

        return png_image_path

    @staticmethod
    def create_aux_folder(folder_name):

        """
        Create auxiliary folder in the directory with executable py-file.
        :param folder_name: string
            Name of folder to be created (not path to it!)
        :return: string
            Absolute path to the created folder.
        """
        # current folder with py-file
        cur_fold = os.path.dirname(os.path.abspath(__file__))
        # path to auxiliary folder
        aux_folder = cur_fold + '/' + folder_name
        # if the folder does not exist yet, create it
        if not os.path.exists(aux_folder):
            os.mkdir(aux_folder)
        # return path to folder
        return aux_folder


class CreateFormMetaData(tk.Frame):

    """
    Create dialog window for user input of metadata tag numbers.

    Attributes
    ----------
    name_of_json_metadata_settings : string
        Relative path to JSON (txt-file) containing string with metadata
        tags' numbers. Used to fill the form with recent settings.
    json_exists_settings : boolean
        Whether JSON (txt-file) containing string with metadata
        tags' numbers already exists.
    text : tkinter's Text object
        Widget containing user input of metadata settings. Is filled
        initially with either values from init_dict if JSON with settings
        does not exist yet, or with recent settings otherwise.
    ok_button_is_created : boolean
        Auxiliary boolean.
        True: Dialog window can be destroyed and program execution continued.

    Methods
    -------
    make_form()
        Create necessary widgets (Label, Text and Button)
        for the dialog window.
    fill_form(value)
        Fill Text widget with value
        (See parameter value of init of class CreateForMetadata).
    read_form()
        Read user input from text widget.
        Save settings to JSON.
        Destroy dialog window (necessary to continue
        the program execution).
    """

    def __init__(self, title, value, master_main_window):

        """

        :param title: string
            Title of dialog window.
        :param value: string or list of hexstrings
            string : Default string of hexnumbers for metadata_tags_list
                defined in init_dict. The variable is of this type if JSON
                with settings does not exist yet.
            list : If the JSON already exists, the settings string of JSON
                is scanned with RegEx and yields list of hexstrings denoting
                needed metadata tag numbers.
        :param master_main_window: tkinter's Tk object
            Master of dialog window.

        """

        # initial arrangements
        super().__init__(master_main_window)
        self.value = value
        self.title = title
        self.master_main_window = master_main_window

        # initialize names for JSONs
        # they will contain settings information
        # and result of program execution (metadata_dictionary)
        self.name_of_json_metadata_settings = 'meta_data_settings.txt'

        # boolean to check if the files already exist
        self.json_exists_settings = os.path.isfile(self.name_of_json_metadata_settings)

        # if JSON with settings already exists, rewrite default value
        if self.json_exists_settings:
            with open(self.name_of_json_metadata_settings, 'r') as file_to_read_info:
                self.value = file_to_read_info.read()

        # auxiliary boolean for destroying main_window
        # i.e. program can be executed after entering tag numbers
        self.ok_button_is_created = False

        # make form of dialog window
        self.make_form()
        # fill the form with default values
        self.fill_form(self.value)

        pass

    def make_form(self):

        """
        Create necessary widgets (Label, Text and Button)
        for the dialog window.

        :return: nothing
        """

        text_for_label = 'Please, enter your desired tags as described bellow:\n\n' \
                         '1) Enter tag pair splitting it by comma\n' \
                         '2) Separate adjacent tag pairs by entering new line\n' \
                         '3) If you wish to reach the subtags\' information,\n' \
                         '   just write parent tag pair followed by child tag pairs\n\n' \
                         'Example:\n' \
                         '        \'0x0040, 0x0275, 0x0032, 0x1060\' (press enter)\n' \
                         '        \'0x7005, 0x1002\' (press enter)\n' \
                         '        \'0x0905, 0x1030\' (press enter)\n' \
                         '         (click on button OK)'

        lab = tk.Label(width=50, text=text_for_label, anchor='w', justify='left')
        lab.pack()
        self.text = tk.Text(self.master_main_window, width=50, height=10)
        self.text.tag_configure("left", justify='left')
        self.text.pack()

        accept_button = tk.Button(master=self.master_main_window, text='OK',
                                  command=self.read_form, anchor='w')
        accept_button.pack()
        self.ok_button_is_created = True

    def fill_form(self, value):

        """
        Fill Text widget with value
        (See parameter value of init of class CreateForMetadata).
        :param value: string or list of hexstrings
            (See parameter value of init of class CreateForMetadata)
        :return: type of variable value
        """

        type_of_value = type(value)
        print(type_of_value)
        if type(value) is list:
            for item in value:
                for sub_item in item:
                    self.text.insert(END, hex(int(sub_item))+', ')
                self.text.insert(END, '\n')
        else:
            list_of_indices = StartClass.recognize_hex_numbers_in_string(settings_string=value)
            list_of_indices = [i for i in list_of_indices if len(i) > 0]
            for prop_index in list_of_indices:
                for tag_element in prop_index:
                    self.text.insert(END, tag_element + ', ')
                self.text.insert(END, '\n')
        return type_of_value

    def read_form(self):

        """
        Read user input from text widget.
        Save settings to JSON.
        Destroy dialog window (necessary to continue
        the program execution).
        :return: string
            (See attribute settings_string of class CreateFormMetadata)
        """

        self.settings_string = self.text.get(1.0, END)
        # write data into JSON

        if self.ok_button_is_created:
            with open(self.name_of_json_metadata_settings, 'w') as json_settings_file:
                json_settings_file.write(self.settings_string)
            self.master_main_window.destroy()

        return self.settings_string


class CreateForm(tk.Frame):

    """
    Create forms for user input.

    - specify exclusion of files from each series folder;
    - selecting size of rectangular ROI when menu 'fixed' is selected;
    - when menu 'array_rois' is selected, specify ROIs array;

    Attributes
    ----------
    entries : list tkinter's objects Entry
        Used to store user's input. Created in method makeform and used in methods
        readform_... for assign contents of entries to variables for further processing.
    accept_button : tkinter's Button object
        For all modes: button to store input information from entries of dialog window.
    entry_width : int
        Length of entry (in symbols) in created form.
    label_width : int
        Length of label (in symbols) in created form.
    color_rect_not_stored : string
        Color of not stored rectangles (blue).
    color_rect_stored : string
        Color of stored rectangles (green).
    color_rect_slided : string
        Color of slided rectangles (red).
    color_rect_resized : string
        Color of resized rectangles (red).
    background : tkinter's Canvas object
        (See attribute background of class GUI)
    file_exclusion_json : string
        Relative path to JSON (txt-file) containing previous settings for file
        exclusion.
    name_of_json : string
        Relative path to JSON (txt-file) containing previous settings for
        the mode 'Fixed_ROI'
    fixed_roi_width : int
        For mode 'Fixed_ROI': width of ROI in pixels.
    fixed_roi_height : int
        For mode 'Fixed_ROI': height of ROI in pixels.
    array : ndarray (2d)
        Pixel array of a dcm-image. Used to get image's size.
        (See attributes px_width and px_height of class CreateForm)
    px_width : int
        Width of dcm-image in pixels. Used for mode 'Array_ROIs' to draw ROIs on Canvas and
        to connect drawn ROIs with pixel array information from DICOMs.
        It is assumed, that all images have the same size.
    px_height : int
        Height of dcm-image in pixels. Used for mode 'Array_ROIs' to draw ROIs on Canvas and
        to connect drawn ROIs with pixel array information from DICOMs.
        It is assumed, that all images have the same size.
    previewed_rect_list : list of tkinter's objects Rectangle
        For mode 'Array_ROIs': shown previewed rectangles, created by clicking on button
        'Preview' in dialog window.
    preview_button : tkinter's Button object
        For mode 'Array_ROIs': button to preview ROIs. Connected with the method
        'preview_array_rois' of class CreateForm.
        Connected to methods beginning with 'readform...' of class CreateForm.
    position_x : int
        For mode 'Array_ROIs': x coordinate (in pixels) of upper left corner of
        upper left ROI in array of ROIs.
    position_y : int
        For mode 'Array_ROIs': y coordinate (in pixels) of upper left corner of
        upper left ROI in array of ROIs.
    number_x : int
        For mode 'Array_ROIs': number of ROIs in array in x direction.
    number_y : int
        For mode 'Array_ROIs': number of ROIs in array in y direction.
    distance_x : int
        For mode 'Array_ROIs': distance in pixels between ROIs in array
        in x direction.
    distance_y : int
        For mode 'Array_ROIs': distance in pixels between ROIs in array
        in y direction.

    Methods
    -------
    makeform()
        Create necessary widgets for the user input form
    fillform(values)
        Fill the form with default or previously specified values
    readform()
        When specifying the size of rectangular ROI ('fixed' menu),
        read the specified values after clicking button 'accept'
    readform_array_roi()
        When specifying parameters of array ROIs,
        read the specified values after clicking button 'accept'
    readform_exclude_files()
        When specifying number of files to be excluded from the beginning
        and the end of each series folder,
        read the specified values after clicking button 'accept'
    preview_array_rois()
        Draw rectangles on the canvas for previewing after clicking
        button 'Preview'
    """

    def __init__(self, master_main_window, fields, values, object_arrays, object_gui, program_start=False):

        """
        :param master_main_window: tkinter's Tk object
            master window for the form's window
        :param fields: list of strings
            list of fields to appear in form
        :param values: list of different data types
            list of the respective values to be filled into fields
        :param object_arrays: instance of the class StartClass
            used to retrieve auxiliary boolean variables and file list (to get images' size)
            it is assumed that all images have same size
        :param object_gui: instance of class GUI
            used to retrieve tool_index (specified mode of ROI selection) and
            to retrieve attributes of this class: colors of the ROI selection's rectangles
        :param program_start: boolean
            auxiliary boolean: 
            if False, attributes from object_arrays and object_gui are retrieved;
            if True, they are not retrieved, because these objects don't exist yet
                (used at program start when file exclusion is to be specified)


        
            """

        # declare attributes
        self.position_x = 0
        self.position_y = 0
        self.number_x = 0
        self.number_y = 0
        self.distance_x = 0
        self.distance_y = 0
        self.entries = []


        self.object_arrays = object_arrays
        self.object_gui = object_gui
        self.program_start = program_start
        self.master_main_window = master_main_window
        super().__init__(master_main_window)
        self.fields = fields
        self.values = values

        if not self.program_start:
            # get measurements of image (assumed they are equal for all images)
            self.array = self.object_arrays.create_base_array(self.object_arrays.filelist[0])['base_array']
            # image measurements
            self.px_height = self.array.shape[0]
            self.px_width = self.array.shape[1]

        # if accept button has been already used, reuse data in the form
        if not self.program_start:
            if self.object_arrays.acceptButtonIsAlreadyUsed:
                self.fixed_roi_width = fixed_roi_width
                self.fixed_roi_height = fixed_roi_height

        # colors of rectangles
        if not self.program_start:
            self.color_rect_not_stored = self.object_gui.color_rect_not_stored
            self.color_rect_stored = self.object_gui.color_rect_stored
            self.color_rect_slided = self.object_gui.color_rect_slided
            self.color_rect_resized = self.object_gui.color_rect_resized

            # initialize source of image shown on GUI
            self.new_source = self.object_gui.new_source
            # retrieve canvas object from the object of GUI
            self.background = self.object_gui.background

            # insert last used parameters of ROI, if they have been already specified
            if self.object_gui.tool_index == 'fixed' and self.object_arrays.acceptButtonIsAlreadyUsed:
                self.values = [self.fixed_roi_height, self.fixed_roi_width]

        # file names of JSONs with stored settings from previous code run sessions
        # for array_rois selection
        self.name_of_json = 'variables_info.txt'
        # for file exclusion
        self.file_exclusion_json = 'file_exclusion_settings.txt'

        # delete all previously previewed rectangles from canvas
        # if obj_gui.previewButtonIsAlreadyUsed:
        #     for rect in self.previewed_rect_list:
        #         obj_gui.background.delete(rect)
        # list for previewed rectangles
        self.previewed_rect_list = []

        # size of labels and entries
        self.label_width = 12
        self.entry_width = 6

        # make form of dialog window
        self.makeform()
        # fill the form with default values
        self.fillform(self.values)

    def makeform(self):

        """
        Create necessary widgets for the user input form.
        :return: nothing
        """

        # list to store frames
        global row_matr
        # list to store entries
        self.entries = []
        # column number of the grid
        t = 0
        # row number of the grid
        q = 0
        row_counter = 0
        row_matr = []
        # iterate through field names
        for field in self.fields:
            # if there more than 50% of fields in one column
            # if q > int(len(self.fields) / 2):
            #     # next column
            #     # +2 since there is also entry
            #     t += 2
            #     # first row
            #     q = 0
            global row
            row = Frame(self.master_main_window)
            lab = Label(row, width=self.label_width, text=field, anchor='w')
            ent = Entry(row, width=self.entry_width)
            row_matr.append(row)

            row.grid(row=q, column=t, padx=5, pady=5)
            lab.grid(row=q, column=t)
            ent.grid(row=q, column=t + 1)
            q += 1
            row_counter += 1
            # store entries as tuple in a list
            self.entries.append((field, ent))

        if not self.program_start:
            if obj_gui.tool_index == 'fixed':
                self.accept_button = tk.Button(master=self.master_main_window, text='accept', command=self.readform, anchor='w')
                self.accept_button.grid(row=len(self.fields) + 1, column=0, padx=5, pady=5)
            elif obj_gui.tool_index == 'array_rois':
                self.accept_button = tk.Button(master=self.master_main_window,
                                               text='accept',
                                               command=self.readform_array_roi,
                                               anchor='w')
                self.accept_button.grid(row=len(self.fields) + 1, column=0, padx=5, pady=5)
                # button for preview
                self.preview_button = tk.Button(master=self.master_main_window,
                                              text='preview',
                                              command=self.preview_array_rois,
                                              anchor='w')

                self.preview_button.grid(row=len(self.fields) + 2, column=0, padx=5, pady=5)

        # if program is just launched, create button for dialog window
        # to exclude some files from the begin and the end of folders with DICOMs
        if self.program_start:
            self.accept_button = tk.Button(master=self.master_main_window, text='continue', command=self.readform_exclude_files, anchor='w')
            self.accept_button.grid(row=len(self.fields) + 1, column=0, padx=5, pady=5)

    def fillform(self, values):

        """
        Fill the form with default or previously specified values

        :param values: list
            Values to fill the form.
        :return: nothing
        """

        counter = 0
        for value in self.values:
            self.entries[counter][1].insert(END, value)
            counter += 1

    def readform(self):

        """
        When specifying the size of rectangular ROI ('fixed' menu),
        read the specified values after clicking button 'accept'

        :return: nothing
        """
        # auxiliary boolean operation to enable storing of last
        # specified parameters of ROI
        self.object_arrays.acceptButtonIsAlreadyUsed = True
        # initialize data_array
        data_array = []
        # fill data_array with values of the form
        for entry in self.entries:
            data_array.append(entry[1].get())
        # pass values to self-variables to use them in the class
        self.fixed_roi_height = data_array[0]
        self.fixed_roi_width = data_array[1]
        # globalize variables to access them in other classes
        global fixed_roi_height
        fixed_roi_height = data_array[0]
        global fixed_roi_width
        fixed_roi_width = data_array[1]
        # destroy dialog window
        self.master_main_window.destroy()

    def readform_array_roi(self):

        """
        When specifying parameters of array ROIs,
        read the specified values after clicking button 'accept'

        :return: nothing
        """

        # delete all previously shown previewed rectangles from canvas
        if obj_gui.previewButtonIsAlreadyUsed:
            for rect in self.previewed_rect_list:
                obj_gui.background.delete(rect)
        # if the button "accept" has been already pressed
        self.object_arrays.acceptButtonIsAlreadyUsed = True
        # array for collecting data from entries
        data_array = []
        # get data from entries
        for entry in self.entries:
            data_array.append(entry[1].get())
        # assignment to self-variables
        self.fixed_roi_height = int(data_array[0])
        self.fixed_roi_width = int(data_array[1])
        self.position_x = int(data_array[2])
        self.position_y = int(data_array[3])
        self.number_x = int(data_array[4])
        self.number_y = int(data_array[5])
        self.distance_x = int(data_array[6])
        self.distance_y = int(data_array[7])
        # self.new_source = obj_gui.new_source
        self.background = obj_gui.background
        # self.px_width = self.object_arrays.im_width_dict[self.new_source]
        # self.px_height = self.object_arrays.im_height_dict[self.new_source]
        # store variables into JSON
        with open(self.name_of_json, 'w') as file_to_store_variables:
            dict_to_store = {'fixed_roi_height': self.fixed_roi_height,
                             'fixed_roi_width': self.fixed_roi_width,
                             'position_x': self.position_x,
                             'position_y': self.position_y,
                             'number_x': self.number_x,
                             'number_y': self.number_y,
                             'distance_x': self.distance_x,
                             'distance_y': self.distance_y,
                             }
            json.dump(dict_to_store, file_to_store_variables)


        # globalize variables to access them in other classes
        global fixed_roi_height
        fixed_roi_height = data_array[0]
        global fixed_roi_width
        fixed_roi_width = data_array[1]

        # build arrays of coordinates

        # then draw rectangles with fixed size
        rect_width = int(self.fixed_roi_width)
        rect_height = int(self.fixed_roi_height)
        # PhX insert pop up window

        l_up_corner_x = int(self.position_x)
        l_up_corner_y = int(self.position_y)

        # initialize position coordinates of link upper corner of first ROI
        x_pos = l_up_corner_x
        y_pos = l_up_corner_y

        # iterate over numbers of ROIs in x-direction (outer loop)
        # and in y-direction (inner loop)
        for x_i in range(self.number_x):
            for y_i in range(self.number_y):
                # specify coordinates of rectangle's diagonal corners
                begin_rect_x = x_pos  # for the scores matrix
                begin_x_im = int(begin_rect_x * obj_gui.background.winfo_width() / self.px_width)  # on the canvas
                end_rect_x = x_pos + rect_width
                end_x_im = int(end_rect_x * obj_gui.background.winfo_width() / self.px_width)
                begin_rect_y = y_pos
                begin_y_im = int(begin_rect_y * obj_gui.background.winfo_height() / self.px_height)
                end_rect_y = y_pos + rect_height
                end_y_im = int(end_rect_y * obj_gui.background.winfo_height() / self.px_height)
                # draw rectangle on canvas
                rect_sel = obj_gui.background.create_rectangle(begin_x_im, begin_y_im, end_x_im, end_y_im,
                                                                 outline=self.color_rect_not_stored)
                # store coordinates of rectangle
                obj_gui.image_rect_coord.append((begin_rect_x,
                                              begin_rect_y,
                                              end_rect_x,
                                              end_rect_y))
                obj_gui.image_rect_coord_record.append((begin_rect_x,
                                                     begin_rect_y,
                                                     end_rect_x,
                                                     end_rect_y))
                obj_gui.rect_coord_dict.update({obj_gui.basename_w_ext: obj_gui.image_rect_coord})
                # store the shown coordinates of rectangles
                obj_gui.image_rect_coord_im.append((begin_x_im,
                                                 begin_y_im,
                                                 end_x_im,
                                                 end_y_im))
                obj_gui.rect_coord_dict_im.update({obj_gui.basename_w_ext: obj_gui.image_rect_coord_im})
                # store the rectangle in the list
                obj_gui.image_rectangles.append(rect_sel)
                obj_gui.rectangles_dict.update({obj_gui.basename_w_ext: obj_gui.image_rectangles})
                # go to the next y_position of ROI
                y_pos += rect_height + self.distance_y
            # go to the next x_position of ROI
            x_pos += rect_width + self.distance_x
            # go to the start y_position of ROI
            y_pos = l_up_corner_y
        # destroy dialog window
        self.master_main_window.destroy()

    def readform_exclude_files(self):

        """
        When specifying number of files to be excluded from the beginning
        and the end of each series folder,
        read the specified values after clicking button 'accept'

        :return: nothing
        """

        # array for collecting data from entries
        data_array = []
        # get data from entries
        for entry in self.entries:
            data_array.append(entry[1].get())
        # assignment to self-variables
        self.num_files_to_exclude_start = int(data_array[0])
        self.num_files_to_exclude_end = int(data_array[1])

        with open(self.file_exclusion_json, 'w') as file_to_store_variables:
            dict_to_store = {'num_files_to_exclude_start': self.num_files_to_exclude_start,
                             'num_files_to_exclude_end': self.num_files_to_exclude_end,
                             }
            json.dump(dict_to_store, file_to_store_variables)

        self.master_main_window.destroy()

    def preview_array_rois(self):

        """
        Draw rectangles on the canvas for previewing after clicking
        button 'Preview'.

        :return:
        """

        # switch auxiliary boolean to True
        obj_gui.previewButtonIsAlreadyUsed = True
        # delete all previously previewed rectangles from canvas
        for rect in self.previewed_rect_list:
            obj_gui.background.delete(rect)
        # array for collecting data from entries
        data_array = []
        # get data from entries
        for entry in self.entries:
            data_array.append(entry[1].get())
        # assignment to self-variables
        self.fixed_roi_height = int(data_array[0])
        self.fixed_roi_width = int(data_array[1])
        self.position_x = int(data_array[2])
        self.position_y = int(data_array[3])
        self.number_x = int(data_array[4])
        self.number_y = int(data_array[5])
        self.distance_x = int(data_array[6])
        self.distance_y = int(data_array[7])


        global fixed_roi_height
        fixed_roi_height = data_array[0]
        global fixed_roi_width
        fixed_roi_width = data_array[1]

        # build arrays of coordinates

        # then draw rectangles with fixed size
        rect_width = int(self.fixed_roi_width)
        rect_height = int(self.fixed_roi_height)
        # PhX insert pop up window

        l_up_corner_x = int(self.position_x)
        l_up_corner_y = int(self.position_y)

        # initialize position coordinates of link upper corner of first ROI
        x_pos = l_up_corner_x
        y_pos = l_up_corner_y
        for x_i in range(self.number_x):
            for y_i in range(self.number_y):
                # specify coordinates of rectangle's diagonal corners
                begin_rect_x = x_pos  # for the scores matrix
                begin_x_im = int(begin_rect_x * obj_gui.background.winfo_width() / self.px_width)  # on the canvas
                end_rect_x = x_pos + rect_width
                end_x_im = int(end_rect_x * obj_gui.background.winfo_width() / self.px_width)
                begin_rect_y = y_pos
                begin_y_im = int(begin_rect_y * obj_gui.background.winfo_height() / self.px_height)
                end_rect_y = y_pos + rect_height
                end_y_im = int(end_rect_y * obj_gui.background.winfo_height() / self.px_height)
                # draw rectangle on canvas
                self.rect_preview = obj_gui.background.create_rectangle(begin_x_im, begin_y_im, end_x_im, end_y_im,
                                                                     outline=self.color_rect_not_stored)
                self.previewed_rect_list.append(self.rect_preview)

                y_pos += rect_height + self.distance_y
            x_pos += rect_width + self.distance_x
            y_pos = l_up_corner_y
        # destroy dialog window
        # self.master_main_window.destroy()


class GUI(tk.Frame):

    """
    Create GUI to show dcm-images and select ROIs.

    Attributes
    ----------
    horiz : int
        Absolute x coordinate of mouse pointer on the image.
        If image has been resized, x position is got as if
        the image has not been resized.
    vert : int
        Absolute y coordinate of mouse pointer on the image.
        If image has been resized, x position is got as if
        the image has not been resized.
    slice_number : int
        Current image number. Got from Slider widget.
    slided_rectangles_list : list of tkinter's Rectangle objects
        If rectangles have been drawn on some image, and then the
        Slider widget was used to slide images, the Rectangle objects
        of previous image are stored in this list.
    slided_rectangles_dict : dict of lists
        Key : current dcm-image file name (not path).
        Value : respective list slided_rectangles_list
            (See the attribute's description in class GUI).
    new_source : string
        Absolute path to current slided dcm-image.
    basename_w_ext : string
        Base name of attribute new_source.
    new_show_im_source : string
        Absolute path to png image respective to currently shown
        dcm-image.
    new_image : PIL.PngImagePlugin.PngImageFile object
        This object is yielded from current png image file
        using Pillow.
    new_img_copy : PIL.PngImagePlugin.PngImageFile object
        Copy of attribute new_image.
        Copy is created to be resized later.
    image : PIL.PngImagePlugin.PngImageFile object
        Attribute new_img_copy resized to actual size of Canvas.
    slice_img : tkinter's PhotoImage object
        This object is finally loaded on canvas widget of main window.
    array : ndarray (2d)
        Pixel array of currently shown dcm-image.
    px_height : int
        Number of rows in attribute array.
    px_width : int
        Number of columns in attribute array.
    slider : tkinter's Slider object
        Slider to slide through images. Connected with method slide_images().
    label_name : tkinter's Label object
        Label containing name of current dcm-image
    label_position : tkinter's Label object
        Label containing absolute x and y position of mouse pointer.
        (See also attributes horiz and vert of class GUI)
    background : tkinter's Canvas object
        Canvas on which the images are displayed.
    menu : tkinter's Menu object
        Menu containing user GUI commands.
        Submenus: Edit, Tools, Final.
    file : tkinter's Menu object
        Submenu of attribute Menu.
        Command: 'Save ROIs and create XLSX'
    image_rect_coord_record : list of tuples of int
        Each tuple for each ROI. Items of each tuple:
            absolute x coordinate of left upper corner of ROI,
            absolute y coordinate of left upper corner of ROI,
            absolute x coordinate of right lower corner of ROI,
            absolute y coordinate of right lower corner of ROI.
    rect_coord_dict : dict of lists of tuples of int
        Keys : base name of dcm-image file
        Values : respective attribute image_rect_coord_record
            (See the attribute's description)
    image_rect_coord_im : list of tuples of int
        Each tuple for each ROI. Items of each tuple:
                shown x coordinate of left upper corner of ROI,
                shown y coordinate of left upper corner of ROI,
                shown x coordinate of right lower corner of ROI,
                shown y coordinate of right lower corner of ROI.
    rect_coord_dict_im : dict of lists of tuples of int
        Keys : base name of dcm-image file
        Values : respective attribute image_rect_coord_im
            (See the attribute's description)
    image_rectangles : list of tkinter's Rectangle objects
        Rectangle objects drawn on current image.
    rectangles_dict : dict of lists of tkinter'S Rectangle objects
        Keys : base name of dcm-image file
        Values : respective attribute image_rectangles
            (See the attribute's description)
    x_coord : int
        Absolute x coordinate of mouse pointer at clicking on
        left mouse button.
    y_coord : int
        Absolute y coordinate of mouse pointer at clicking on
        left mouse button.
    all_roi_dict : dict of lists of tuples
        Keys : absolute paths to dcm-images
        Values : lists of tuples containing coordinates of
            left upper (x and y) and right lower (x and y)
            corners of ROIs. Each tuple for each ROI.
            (See attribute image_rect_coord_record)
    form : instance of class CreateForm
        Object containing information needed in mode 'Fixed_ROIs'
        about size of fixed ROI.

    Methods
    -------
    motion(self, event)
        Get absolute mouse pointer coordinates, when moving it.
        Update mouse pointer position label's text.
    slide_images(self, event)
        Slide shown images with Slider widget.
    create_widgets(self)
        Create necessary widgets (Labels, Canvas, Slider, Menu)
        for main window.
    _resize_image(self, event)
        Adjust shown image size, canvas size and
        slider width to resized main window.
    create_bindings(self)
        Connect events to methods.
    draw(self, event)
        Select ROIs depending on selecting mode:
        'Fixed_ROI' or 'Draw_Diagonal'.
        Draw rectangles on canvas.
        Store their absolute coordinates in self.image_rect_coord.
        Store their resized coordinates in self.image_rect_coord_im.
        Store Rectangle objects in self.rectangles_dict.
    erase_last(self)
        Erase last ROI drawn on canvas.
    update_roi_dict(self)
        Update attribute all_roi_dict, when choosing
        menu 'Save ROIs and create XLSX-file'.
    choose_whole_image(self)
        Select whole images as ROIs and store their pixel arrays
        into all_roi_dict.
    minimal_shape(self, dict_of_2d_arrays)
        minimal_shape(self, dict_of_2d_arrays).
    choose_fixed(self)
        Select 'Fixed_ROIs' mode.
        Create form for choosing fixed ROI's size.
    choose_array_rois(self)
        Select 'Array_ROIs' mode.
        Create form for choosing parameters of array of ROIs.
    choose_diagonal(self)
        Select mode 'Draw_Diagonal'
    build_initial_coord(self, event)
        Store absolute coordinates of mouse pointer
        when clicking on left mouse button.
    """

    def __init__(self, *, master, obj_arrays, crop_perc, fit_order, useFitting, im_height_in_mm,
                 im_width_in_mm, extensions, trunc_percentage,
                 useCentralCropping, start_freq_range, end_freq_range, step,
                 useTruncation, multipleFiles, pixel_size_in_mm, equal_ROIs_for_each_image,
                 main_window_width, main_window_height):

        """

        :param master: tkinter's Tk object
            Master of window with shown images.
        :param obj_arrays: instance of class StartClass
            Used to retrieve its attributes: acceptButtonIsAlreadyUsed,
            filelist, filedict, all_images; and method create_base_array().
        :param crop_perc:
        :param fit_order: int (1 or 2)
            If attribute useFitting is True, the order of fitting of dcm-image.
            Specified in init_dict.
        :param useFitting:
            True: 1st or 2nd order fit of dcm images is executed as a method of
                background removal.
            False: Mean value subtraction is used as background removal.
            This attribute is specified in init_dict.
        :param im_height_in_mm: float
            Height of FOV (Field-Of-View) in mm. Specified in init_dict.
        :param im_width_in_mm:
            Width of FOV in mm. Specified in init_dict.
        :param extensions: list of strings
            List of extensions to be searched for in selected directory.
            Specified in init_dict.
        :param trunc_percentage: float
            Under this per cent value the NPS is truncated at
            the end of NPS-values list. Specified in init_dict.
        :param x0: int
            x-coordinate of mouse pointer when clicking on left mouse
            button.
        :param x1: int
            x-coordinate of mouse pointer when releasing left mouse
            button.
        :param y0:
            y-coordinate of mouse pointer when clicking on left mouse
            button.
        :param y1:
            y-coordinate of mouse pointer when releasing left mouse
            button.
        :param useCentralCropping: boolean
            Whether the ROI should be a central crop of the dcm-images.
            True: central part of the image is cropped.
                Percentage of the cropping is specified by attribute
                crop_percentage of class GUI.
            Specified in init_dict.
        :param start_freq_range: float
            Start of frequency range (maybe not needed in this class)
        :param end_freq_range: float
            End of frequency range (maybe not needed in this class)
        :param step: float
            Step of frequency range (maybe not needed in this class)
        :param useTruncation: boolean
            Whether truncation of low valued NPS at higher frequencies
            should be used. Specified in init_dict.
        :param multipleFiles: (not needed in this class)
        :param pixel_size_in_mm: float
            Default pixel size in mm. (If metadata PixelSpacing
            is not known). Specified in init_dict.
        :param equal_ROIs_for_each_image: boolean
            Whether there should be equal ROIs selected in all images.
            Specified in init_dict.
        :param main_window_width: int
            Width of main window (showing images, labels, slider, etc.)
            in pixels. Specified in init_dict.
        :param main_window_height: int
            Height of main window (showing images, labels, slider, etc.)
            in pixels. Specified in init_dict.
        """

        print('Constructor of the class GUI is being executed')


        self.master = master
        super().__init__(master)  # master frame
        self.grid()  # packing of master frame on grid
        # get base arrays as dict from another class
        self.obj_arrays = obj_arrays

        # initialize attributes

        # defined in self.motion()
        self.horiz = 0
        self.vert = 0
        # defined in self.slide_images()
        self.slice_number = 0
        self.basename_w_ext = ''
        self.new_image = ''
        self.new_img_copy = ''
        self.image = ''
        self.slice_img = ''
        # defined in create_widgets()
        self.slider = 0
        self.label_name = 0
        self.label_position = 0
        self.background = 0
        # defined in choose_fixed()
        self.form = 0
        self.dialog_window = 0











        self.freq = []
        # self.arrays_dict = self.obj_arrays.arrays_dict
        # slider's height
        self.slider_height = 15
        # height of label
        self.label_height = 57
        # aux boolean
        self.already_resized = False
        self.counter_pass = 0
        # color of rectangles (not stored)
        self.color_rect_not_stored = '#0000FF'
        # color of rectangles (not stored)
        self.color_rect_slided = '#FF0000'
        # color of rectangles (stored)
        self.color_rect_stored = '#00FF00'
        # color of resized rectangles
        self.color_rect_resized = '#0000FF'
        # color of active button
        self.color_button_active = '#9FFF80'
        # color of inactive button
        self.color_button_inactive = '#E6E6E6'
        # field names and initial values for dialog frame
        self.fields = ['height_px', 'width_px']
        # size of main window
        self.main_window_height = main_window_height
        self.main_window_width = main_window_width
        self.size_of_main_window = max(self.main_window_width, self.main_window_height)
        if self.obj_arrays.acceptButtonIsAlreadyUsed:
            self.values = [int(fixed_roi_height), int(fixed_roi_width)]
        else:
            self.values = [128, 128]
        # field names for array rois selection
        self.fields_array_rois = ['height',
                                  'width',
                                  'left_upper_x',
                                  'left_upper_y',
                                  'number_roi_x',
                                  'number_roi_y',
                                  'distance_x',
                                  'dictance_y']
        if os.path.isfile('variables_info.txt'):
            with open('variables_info.txt', 'r') as file_to_read_info:
                read_dict = json.load(file_to_read_info)
            list_of_variables = [read_dict[key] for key in read_dict]
            self.values_array_rois = list_of_variables
        else:
            self.values_array_rois = [64,
                                      64,
                                      10,
                                      10,
                                      1,
                                      1,
                                      5,
                                      5]
        # auxuliary boolean to destroy perviewed rectangles
        self.previewButtonIsAlreadyUsed = False
        # size of dialog window
        self.dialog_window_height = 110
        self.dialog_window_width = 160
        self.dialog_window_x = 200
        self.dialog_window_y = 20

        # for array of ROIs
        self.dialog_window_height_arr_roi = 400
        self.dialog_window_width_arr_roi = 190
        self.dialog_window_x_arr_roi = 200
        self.dialog_window_y_arr_roi = 20
        # pixel size
        self.pixel_size_in_mm = pixel_size_in_mm
        # whether there are equal ROIs for each image
        self.equalROIs = equal_ROIs_for_each_image
        # whether it should be one or multiple excel-files
        self.multipleFiles = multipleFiles
        # all rectangles
        self.all_rect = []
        # dictionary for rectangles on each image
        self.all_rect_dict = {}
        # dictionary for rectangle coordinates
        self.rect_coord_dict = {}
        self.rect_coord_dict_im = {}
        self.image_rect_coord = []
        self.image_rect_coord_record = []
        self.image_rect_coord_im = []
        # dictionary for rectangles in slided images
        self.slided_rectangles_dict = {}
        # list of slided rectangles in current image
        self.slided_rectangles_list = []
        # list for rectangles created by resizing

        # dictionary for rectangle objects on each image
        self.rectangles_dict = {}
        self.image_rectangles = []
        # ROIs for current image
        self.image_roi = []
        # dictionary for ROIs of all images
        self.all_roi_dict = {}
        # collect all recorded rectangles
        self.all_rect_record = []
        # all xlsx-files
        self.all_xlsx_range = []
        self.all_xlsx = []
        # paths to all cropped images
        self.all_cropped_im = []
        # paths to all one d nps images
        self.all_1_d_nps = []
        # whether central cropping should be applied
        self.useCentralCropping = useCentralCropping
        # whether lower nps should be deleted
        self.useTruncation = useTruncation
        # extensions of image files
        self.extensions = extensions
        # image measurements in mm
        # self.im_height_in_mm = im_height_in_mm
        # self.im_width_in_mm = im_width_in_mm
        # truncate lower nps
        self.trunc_percentage = trunc_percentage
        # maximal size of the image (height or width)
        # self.max_size = max(im_height_in_mm, im_width_in_mm)
        # cropping percentage
        self.crop_perc = crop_perc
        # fitting order for 2d-fit
        self.fit_order = fit_order
        # whether fitting should be applied
        # or background removal should be used
        self.useFitting = useFitting

        # freq range parameters
        self.start_freq = start_freq_range
        self.end_freq = end_freq_range
        self.num_of_steps = int((self.end_freq - self.start_freq) // step + 1)
        self.freq_range = np.linspace(start=self.start_freq,
                                      stop=self.end_freq,
                                      num=self.num_of_steps)

        # get the duration of program's execution
        self.start_timer = time.time()
        # create list of image files (paths to them)
        self.file_list = self.obj_arrays.filelist
        self.filedict = self.obj_arrays.filedict
        # initialize empty dictionary for rois
        print('creation of all_roi_dict is being executed')
        for path_to_file in self.file_list:
            key = path_to_file
            self.all_roi_dict.update({key: []})
        print('creation of all_roi_dict is done')


        # initial image array
        self.array = self.obj_arrays.create_base_array(self.file_list[0])['base_array']
        # initial label name
        self.new_source = self.file_list[0]
        # initial image for canvas
        self.new_show_im_source = self.obj_arrays.all_images[0]
        # initial image height and width
        # image measurements
        self.px_height = self.array.shape[0]
        self.px_width = self.array.shape[1]
        self.new_im_size = max(self.px_width, self.px_height)
        self.resize_coef = 1
        # initial tool_index
        self.tool_index = 'diagonal'
        # create widgets
        self.create_widgets()
        # create bindings
        self.create_bindings()

        print('Constructor of the class GUI is done')


    def motion(self, event):
        """
        Get absolute mouse pointer coordinates, when moving it.
        Update mouse pointer position label's text.

        If the main window is resized self.vert and self.horiz
        define absolute position of mouse pointer as if the image was not resized.

        :param event: tkinter's event object
            Mouse event Motion.
        :return: nothing
        """
        try:
            # horiz. mouse pointer position in the changed frame
            self.horiz = int(self.px_width * event.x / self.background.winfo_width())
            # vert. mouse pointer position in the changed frame
            self.vert = int(self.px_height * event.y / self.background.winfo_height())
            # update label's text
            self.label_position["text"] = 'x: ' + str(self.horiz) + '| y: ' + str(self.vert)
        except NameError:
            pass
        except AttributeError:
            pass

    def slide_images(self, event):

        """
        Slide shown images with Slider widget.
        :param event: tkinter's event object
            Slider Movement object.
        :return: nothing
        """
        # clear the list of slided rectangles
        self.slided_rectangles_list = []
        # get image number from slider
        self.slice_number = self.slider.get() - 1
        # path to file containing pixel data
        self.new_source = self.file_list[self.slice_number]
        self.basename_w_ext = os.path.basename(self.new_source)
        # path to new png image
        self.new_show_im_source = self.obj_arrays.all_images[self.slice_number]
        # create Pillow image object
        self.new_image = Image.open(self.new_show_im_source)
        # copy object of new image
        self.new_img_copy = self.new_image.copy()
        # resize the copy to new size specified by user
        self.image = self.new_img_copy.resize((self.new_im_size, self.new_im_size))
        # create PhotoImage object for canvas (self.background) from the resized image
        self.background.new_background_image = ImageTk.PhotoImage(self.image)
        # load the object PhotoImage on the Canvas
        self.slice_img = self.background.create_image(0, 0, anchor=NW, image=self.background.new_background_image)
        # draw rectangles selected in previous images
        # and store the respective Rectangle objects in a list
        try:
            for coord in self.image_rect_coord_im:
                self.slided_rectangles_list.append(
                    self.background.create_rectangle(
                        coord[0],  # * self.resize_coef,
                        coord[1],  # * self.resize_coef,
                        coord[2],  # * self.resize_coef,
                        coord[3],  # * self.resize_coef,
                        outline=self.color_rect_slided))
        except KeyError:
            pass

        # update dict of slided rectangles
        self.slided_rectangles_dict.update({self.basename_w_ext: self.slided_rectangles_list})
        # print('slided_rectangles:  ', self.slided_rectangles_dict)

        # change title of main window according to new image file
        self.master.title(os.path.basename(self.new_source))
        # create array for current image
        self.array = self.obj_arrays.create_base_array(self.new_source)['base_array']

        # image measurements
        self.px_height = self.array.shape[0]
        self.px_width = self.array.shape[1]

        # self.px_width = self.obj_arrays.im_width_dict[self.new_source]
        # self.px_height = self.obj_arrays.im_height_dict[self.new_source]
        # change label name

        # update name label's text
        self.label_name["text"] = os.path.basename(self.new_source)

    def create_widgets(self):
        """
        Create necessary widgets (Labels, Canvas, Slider, Menu) 
        for main window. 
        :return: nothing
        """
        # slider
        self.slider = Scale(master=self.master, from_=1, to=2, length=self.px_width,
                            orient=HORIZONTAL, highlightthickness=0, bd=0, width=self.slider_height,
                            command=self.slide_images)
        self.slider.grid(row=4)

        # label for base name of current image file
        self.label_name = tk.Label(self, highlightthickness=0, bd=0)  # create the demonstration label
        self.label_name.grid(row=1, column=0)

        # label for position of mouse pointer on the image
        self.label_position = tk.Label(self, highlightthickness=0, bd=0)  # create the demonstration label
        self.label_position.grid(row=2, column=0)
        # passing the initial text
        self.label_name["text"] = os.path.basename(self.new_source)
        self.label_position["text"] = 'set pointer on the image'

        # canvas with image of slice
        self.background = Canvas(self, highlightthickness=0, borderwidth=0, width=self.px_width,
                                 height=self.px_height)
        self.background.grid(row=3)
        imp = Image.open(self.new_show_im_source)
        self.background.image = ImageTk.PhotoImage(imp)
        self.slice_img = self.background.create_image(0, 0, anchor=NW, image=self.background.image)

        # create menubar
        self.menu = tk.Menu(self.master)
        self.master.config(menu=self.menu)

        # add file menu
        self.file = tk.Menu(self.menu)
        self.file.add_command(label='Save ROIs and create XLSX-file', command=self.update_roi_dict)
        # self.file.add_command(label='Create XLSX', command=self.command_for_create_xlsx_button)
        # self.menu.add_cascade(label='File', menu=self.file)

        # add edit menu
        edit = tk.Menu(self.menu)
        edit.add_command(label='Erase Last ROI', command=self.erase_last)
        self.menu.add_cascade(label='Edit', menu=edit)

        # add tools menu
        tools = tk.Menu(self.menu)
        tools.add_command(label='Fixed Rectangle', command=self.choose_fixed)
        tools.add_command(label='Draw Diagonal', command=self.choose_diagonal)
        tools.add_command(label='Array of ROIs', command=self.choose_array_rois)
        tools.add_command(label='Whole Images', command=self.choose_whole_image)
        self.menu.add_cascade(label='Tools', menu=tools)

    def _resize_image(self, event):
        """
        Adjust shown image size, canvas size and 
        slider width to resized main window.
        :param event: tkinter's event object
            Configure event.
        :return: nothing
        """
        # store previous size of window
        if self.already_resized:
            self.previous_size = min(self.new_height, self.new_width)
        # get new width and height of window
        self.new_width = self.master.winfo_width()  # actual width of main window
        self.new_height = self.master.winfo_height()  # actual height of main window
        # condition that checks whether window has been resized
        if self.already_resized:
            if np.abs(self.previous_size - min(self.new_height, self.new_width)) < 5:
                # print('pass %d' % self.counter_pass)
                self.counter_pass += 1
                pass
        self.already_resized = True
        # new size of the image
        if self.new_width < self.new_height - (self.slider_height + self.label_height):
            self.new_im_size = self.new_width
        else:
            self.new_im_size = self.new_height - (self.slider_height + self.label_height)
        # resizing coefficient
        self.resize_coef = min(self.new_height, self.new_width) / self.size_of_main_window

        self.size_of_main_window = min(self.new_height, self.new_width)
        self.image_heigth = self.background.winfo_height()  # actual height of slice image
        # position of slider
        self.z = self.slider.get() - 1
        # source of the shown image
        self.new_show_im_source = self.obj_arrays.all_images[self.z - 1]
        # open image with PIL
        self.new_image = Image.open(self.new_show_im_source)
        # make copy of the opened image
        self.new_img_copy = self.new_image.copy()
        # resize image copy according to actual main window size
        self.image = self.new_img_copy.resize((self.new_im_size, self.new_im_size))
        # put the resized image on canvas
        self.background.new_background_image = ImageTk.PhotoImage(self.image)
        self.background.itemconfig(self.slice_img, image=self.background.new_background_image)
        # resize canvas
        self.background.configure(width=self.new_im_size)
        self.background.configure(height=self.new_im_size)
        # resize the length of the slider
        self.slider['length'] = self.new_width

    def create_bindings(self):

        """
        Connect events to methods.
        :return: nothing.
        """
        # resize image method
        self.master.bind('<Configure>', self._resize_image)
        # click on left mouse button
        self.background.bind('<Button-1>', self.build_initial_coord)
        # release of left mouse button
        self.background.bind('<ButtonRelease-1>', self.draw)
        # bind motion event to display coordinates of mouse pointer
        self.background.bind('<Motion>', self.motion)

    def draw(self, event):
        """
        Select ROIs depending on selecting mode:
        'Fixed_ROI' or 'Draw_Diagonal'.
        Draw rectangles on canvas.
        Store their absolute coordinates in self.image_rect_coord.
        Store their resized coordinates in self.image_rect_coord_im.
        Store Rectangle objects in self.rectangles_dict.
        :param event: tkinter's event object
            <'ButtonRelease-1'> event.
        :return: nothing.
        """
        # the words 'absolute size, height, width etc.' mean size of ROI, image etc.
        # before resizing
        if self.tool_index == 'fixed':
            # then draw rectangles with fixed size (absolute size)
            rect_width = int(self.form.fixed_roi_width)  # absolute width of rectangle
            rect_height = int(self.form.fixed_roi_height)  # absolute height of rectangle
            # define coordinates of the left upper corner of rectangle (absolute values)
            l_up_corner_x = int(event.x * self.px_width / self.background.winfo_width())
            l_up_corner_y = int(event.y * self.px_height / self.background.winfo_height())
            # specify coordinates of rectangle's diagonal corners
            begin_rect_x = l_up_corner_x  # for the scores matrix
            begin_x_im = int(begin_rect_x * self.background.winfo_width() / self.px_width)  # on the canvas
            end_rect_x = l_up_corner_x + rect_width
            end_x_im = int(end_rect_x * self.background.winfo_width() / self.px_width)
            begin_rect_y = l_up_corner_y
            begin_y_im = int(begin_rect_y * self.background.winfo_height() / self.px_height)
            end_rect_y = l_up_corner_y + rect_height
            end_y_im = int(end_rect_y * self.background.winfo_height() / self.px_height)
            # draw rectangle on canvas
            rect_sel = self.background.create_rectangle(begin_x_im, begin_y_im, end_x_im, end_y_im,
                                                             outline=self.color_rect_not_stored)
            # store absolute coordinates of rectangle
            self.image_rect_coord.append((begin_rect_x,
                                          begin_rect_y,
                                          end_rect_x,
                                          end_rect_y))
            self.image_rect_coord_record.append((begin_rect_x,
                                          begin_rect_y,
                                          end_rect_x,
                                          end_rect_y))
            self.rect_coord_dict.update({self.basename_w_ext: self.image_rect_coord})
            # store the shown coordinates of rectangles
            self.image_rect_coord_im.append((begin_x_im,
                                          begin_y_im,
                                          end_x_im,
                                          end_y_im))
            self.rect_coord_dict_im.update({self.basename_w_ext: self.image_rect_coord_im})
            # store the rectangle in the list
            self.image_rectangles.append(rect_sel)
            self.rectangles_dict.update({self.basename_w_ext: self.image_rectangles})
            # subarray for roi
            # for key, path_to_image in zip(self.all_roi_dict, self.arrays_dict):
            #     self.current_array = self.arrays_dict[path_to_image]
            #     self.subarray_roi = self.current_array[l_up_corner_y : l_up_corner_y + rect_height,
            #                         l_up_corner_x: l_up_corner_x + rect_width]
            #     image_roi = self.all_roi_dict[key]
            #     image_roi.append(self.subarray_roi)
            #     print('shape_of_image_roi', np.array(image_roi).shape)
            #     self.all_roi_dict.update({key: image_roi})
                # image_roi = []
            # delete event coordinates
            del self.x_coord
            del self.y_coord
            # print('shape_of_image_roi_total', np.array(image_roi).shape)

        elif self.tool_index == 'diagonal':
            # coordinates are normalized to window size
            self.x_coord.append(event.x * self.px_width / self.background.winfo_width())
            self.y_coord.append(event.y * self.px_height / self.background.winfo_height())

            # get current slice
            self.z = self.slider.get() - 1  # "-1" because the first slice has number zero

            # specify coordinates of rectangle's diagonal corners
            begin_rect_x = int(self.x_coord[0])  # for the scores matrix
            begin_x_im = int(begin_rect_x * self.background.winfo_width() / self.px_width)  # on the canvas
            end_rect_x = int(self.x_coord[1])
            end_x_im = int(end_rect_x * self.background.winfo_width() / self.px_width)
            begin_rect_y = int(self.y_coord[0])
            begin_y_im = int(begin_rect_y * self.background.winfo_height() / self.px_height)
            end_rect_y = int(self.y_coord[1])
            end_y_im = int(end_rect_y * self.background.winfo_height() / self.px_height)
            # draw rectangle on canvas
            rect_sel = self.background.create_rectangle(begin_x_im, begin_y_im, end_x_im, end_y_im,
                                                             outline=self.color_rect_not_stored)
            # store the rectangle in the list
            self.image_rectangles.append(rect_sel)
            self.rectangles_dict.update({self.basename_w_ext: self.image_rectangles})
            # make possible to draw rectangle from all directions
            if begin_rect_x < end_rect_x:
                x0 = begin_rect_x
                x0_im = begin_x_im
                x1 = end_rect_x
                x1_im = end_x_im
            else:
                x1 = begin_rect_x
                x1_im = begin_x_im
                x0 = end_rect_x
                x0_im = end_x_im

            if begin_rect_y < end_rect_y:
                y0 = begin_rect_y
                y0_im = begin_y_im
                y1 = end_rect_y
                y1_im = end_y_im
            else:
                y1 = begin_rect_y
                y1_im = begin_y_im
                y0 = end_rect_y
                y0_im = end_y_im
            # set minimal size of ROI 5x5px
            if (x1 - x0) < 5:
                x1 = x0 + 5
            if (y1 - y0) < 5:
                y1 = y0 + 5
            self.subarray_roi = self.array[y0:y1, x0:x1]
            if len(self.subarray_roi) == 0:  # DEBUG
                print('x0: ', x0)
                print('x1: ', x1)
                print('y0: ', y0)
                print('y1: ', y1)
            # delete event coordinates
            del self.x_coord
            del self.y_coord
            # store coordinates of rectangle
            self.image_rect_coord.append((x0,
                                          y0,
                                          x1,
                                          y1))
            self.image_rect_coord_record.append((x0,
                                                 y0,
                                                 x1,
                                                 y1))
            self.rect_coord_dict.update({self.basename_w_ext: self.image_rect_coord})
            # store the shown coordinates of rectangles
            self.image_rect_coord_im.append((x0_im,
                                             y0_im,
                                             x1_im,
                                             y1_im))
            self.rect_coord_dict_im.update({self.basename_w_ext: self.image_rect_coord_im})
            # store the rectangle in the list
            self.image_rectangles.append(rect_sel)
            self.rectangles_dict.update({self.basename_w_ext: self.image_rectangles})
        # elif self.tool_index == 'array_roi':

    def erase_last(self):
        """
        Erase last ROI drawn on canvas.
        :return: nothing.
        """
        print('Last ROI has been deleted')
        try:
            # delete the last drawn rectangle from canvas
            self.background.delete(self.image_rectangles[-1])
            del self.image_rectangles[-1]
        except IndexError:
            pass
        try:
            # delete rectangle if it has been drawn on another image
            # and is shown as slided rectangle
            self.background.delete(self.slided_rectangles_list[-1])
            # delete last item of slided_rectangles_list
            del self.slided_rectangles_list[-1]
        except IndexError:
            pass
        try:
            # delete last roi coordinates from roi dictionary
            for key in self.all_roi_dict:
                roi_arr = self.all_roi_dict[key]
                del roi_arr[-1]
                self.all_roi_dict.update({key: roi_arr})
            print('im_roi:    ', np.array(roi_arr).shape)  # DEBUG
        except IndexError:
            pass

        try:
            # delete coordinates of rectangles
            del self.image_rect_coord_im[-1]
        except IndexError:
            pass

        try:
            del self.image_rect_coord[-1]
        except IndexError:
            pass
        print('%d ROIs remain' % (len(self.image_rectangles)))  # DEBUG

    def store_ROIs(self, *args):
        self.image_roi.append(self.subarray_roi)
        print(np.array(self.image_roi).shape)

    def update_roi_dict(self):
        """
        Update attribute all_roi_dict, when choosing
        menu 'Save ROIs and create XLSX-file'.
        :return: nothing.
        """
        print('ROIs are being saved. It can take some minutes')
        for coord in self.image_rect_coord:
            for key, path_to_file in zip(self.all_roi_dict, self.file_list):
                roi_arr = self.all_roi_dict[key]
                roi_arr.append(coord)
                self.all_roi_dict.update({key: roi_arr})
        for rect in self.image_rectangles:
            self.background.itemconfig(rect, outline=self.color_rect_stored)
        # change the color of resized rectangles
        self.color_rect_resized = '#00FF00'
        # clear rect coord array to not duplicate ROIs
        self.image_rect_coord = []
        # print(self.all_roi_dict)
        print('ROIs have been saved')
        global roi_processing_start_time
        roi_processing_start_time = time.time()
        obj_process_roi.execute_calc_nps_sorted()

    def choose_whole_image(self):
        """
        Select whole images as ROIs and store their pixel arrays
        into all_roi_dict.
        :return: nothing
        """
        # get minimal shape of images
        # minimal_shape = self.minimal_shape(dict_of_2d_arrays=self.arrays_dict)
        # min_rows = minimal_shape['rows']
        # min_columns = minimal_shape['columns']
        # iterate over all images
        for key_path in self.file_list:
            # update all_roi_dict
            key_image = os.path.basename(key_path)
            # self.all_roi_dict.update({key_image: np.array([
            #     self.arrays_dict[key_path][:min_rows][:min_columns]])})
            self.all_roi_dict.update({key_image: np.array([
                self.obj_arrays.create_base_array(key_path)['base_array']])})

    def minimal_shape(self, dict_of_2d_arrays):
        """
        Return minimal shape of 2d_arrays.
        :param dict_of_2d_arrays: dict of nd_arrays (2d)
        :return: dict with keys 'columns' and 'rows'
        """
        # initialize minimal shape
        min_rows = 10**10
        min_columns = 10**10
        # iterate over all keys of dictionary
        for key in dict_of_2d_arrays:
            # get 2d array in variable
            array_2d = dict_of_2d_arrays[key]
            num_of_columns = array_2d.shape[1]
            num_of_rows = array_2d.shape[0]
            # update minimal shape values
            if min_columns > num_of_columns:
                min_columns = num_of_columns
            if min_rows > num_of_rows:
                min_rows = num_of_rows
        return_dictionary = {'columns': min_columns,
                             'rows': min_rows}
        return return_dictionary

    def choose_fixed(self):
        """
        Select 'Fixed_ROIs' mode.
        Create form for choosing fixed ROI's size.
        :return: nothing
        """
        self.tool_index = 'fixed'
        self.dialog_window = tk.Toplevel(master=self.master)
        self.dialog_window.geometry('%dx%d+%d+%d' % (self.dialog_window_width,
                                                     self.dialog_window_height,
                                                     self.dialog_window_x,
                                                     self.dialog_window_y))
        self.dialog_window.title('Initial Data')
        self.form = CreateForm(master_main_window=self.dialog_window,
                               fields=self.fields,
                               values=self.values,
                               object_arrays=self.obj_arrays,
                               object_gui=obj_gui)
        pass

    def choose_array_rois(self):
        """
        Select 'Array_ROIs' mode.
        Create form for choosing parameters of array of ROIs.
        :return: nothing
        """
        self.tool_index = 'array_rois'
        self.dialog_window = tk.Toplevel(master=self.master)
        self.dialog_window.geometry('%dx%d+%d+%d' % (self.dialog_window_width_arr_roi,
                                                     self.dialog_window_height_arr_roi,
                                                     self.dialog_window_x_arr_roi,
                                                     self.dialog_window_y_arr_roi))
        self.dialog_window.title('Initial Data For Array Of ROIs')
        # if the variables have already been stored in JSON-file, read them from it
        if os.path.isfile('variables_info.txt'):
            with open('variables_info.txt', 'r') as file_to_read_info:
                read_dict = json.load(file_to_read_info)
            list_of_variables = [read_dict[key] for key in read_dict]
            self.values_array_rois = list_of_variables
        self.form_array_roi = CreateForm(master_main_window=self.dialog_window,
                               fields=self.fields_array_rois,
                               values=self.values_array_rois,
                               object_arrays=self.obj_arrays,
                               object_gui=obj_gui)
        pass

    def choose_diagonal(self):
        """
        Select mode 'Draw_Diagonal'
        :return: nothing
        """
        self.tool_index = 'diagonal'

    def build_initial_coord(self, event):
        """
        Store absolute coordinates of mouse pointer
        when clicking on left mouse button.
        :param event: tkinter's event object
             <'Button-1'> event.
        :return: nothing
        """
        self.x_coord = []
        self.y_coord = []
        self.x_coord.append(event.x * self.px_width / self.background.winfo_width())
        self.y_coord.append(event.y * self.px_height / self.background.winfo_height())
        return

    @staticmethod
    def create_aux_folder(folder_name, cur_fold=None):

        """
        Create auxiliary folder in the directory with executable py-file.

        :param folder_name: string
            Name of folder to be created (not path to it!)
        :param cur_fold: string
            Path to initial folder, where the new folder is to be created.
            If None (by default), the new folder is created in the same
            directory with this py-file.
        :return: string
            Absolute path to the created (or already existing) folder.
        """

        # current folder with py-file
        if cur_fold is None:
            cur_fold = os.path.dirname(os.path.abspath(__file__))
        # path to auxiliary folder
        aux_folder = cur_fold + '/' + folder_name
        # if the folder does not exist yet, create it
        if not os.path.exists(aux_folder):
            os.mkdir(aux_folder)
        # return path to folder
        return aux_folder


class ProcessROI:

    """
    Calculate NPS and create files with results.

    Attributes
    ----------
    workbook_series : XlsxWriter's Workbook object
        Workbook containing NPS info for each image in current series.
        Each worksheet (except the last one) contains NPS info for current dcm-image.
        Last worksheet contains averaged NPS info ansd is called 'averaged'.
        Near description is to find in the manual.
    workbook_averaged : XlsxWriter's Workbook object
        Workbook containing all averaged worksheets of attributes workbook_series
        for current study.
        I.e. each worksheet contains averaged NPS info for current series.
    workbook_summary : openpyxl's Workbook object
        Workbook containing summarized information of averaged NPS xlsx-files.
    worksheet_averaged : XlsxWriter's Worksheet object
        Worksheet of attribute workbook_averaged.
    worksheet_summary : openpyxl'S Worksheet object
        Worksheet of attribute workbook_summary.
    col_number, col_folder, col_series, col_peak_freq,
    col_peak_value, col_left_dev, col_right_dev,
    col_area, col_ave_m_HU : strings
        Excel letters for each information column in workbook_summary.
    metadata_columns : list of strings
        List of Excel letetrs for columns containing metadata info in
        workbook_summary.
    all_roi_dict : dict of dict of lists of tuples
        (See attribute all_roi_dict of class GUI)
    sorted_all_roi_dict : dict of dicts
        Dict with same structure as attribute filedict of class StartClass.
        The lists of files in series folder are transformed into dictionaries:
        Keys : absolute path to image in current series folder
        Values : list of tuples with ROIs' diagonal coordinates:
            x0 - x coordinate of left upper corner
            y0 - y coordinate of left upper corner
            x1 - x coordinate of right lower corner
            y1 - y coordinate of right lower corner
    image_rect_coord : list
        (See attribute image_rect_coord of class GUI)
    image_rect_coord_record : list
        (See attribute image_rect_coord_record of class GUI)
    roi_image_mean_HU : list of floats
        List of mean HU for each ROI in current image.
    image_sd : list of floats
        List of standard deviation for each ROI in current image.
    all_mean_HU_dict : dict of lists of floats
        Dict of attributes roi_image_mean_HU for current series folder.
        Keys : paths to dcm-image file.
        Values : respective attribute roi_image_mean_HU
            (See description of roi_image_mean_HU above)
    all_SD_dict : dict of lists of floats
        Dict of attributes image_sd for current series folder.
        Keys : paths to dcm-image file.
        Values : respective attribute image_sd
            (See description of image_sd above)        
    nps_image : list of dicts
        For current image.
        List of range_dict attributes for each ROI
        on current image.
        For each ROI on image, range_dict consists of:
            'values': interpolated NPS list of the ROI
            'frequencies': respective frequencies
            'AUC': area under not interpolated 1d-NPS profile
            'integral_of_2d_nps': integral of 2d NPS of teh ROI
                                    (not interpolated!!!)
    all_nps_dict : dict_ of lists of dicts
        For images in current series folder.
        Keys : paths to images in current series folder
        Values : nps_image attributes for respective image
    all_average_nps : dict of dicts
        For images in current series folder.
        Keys : absolute paths to images in current series folder.
            Values : dict
                Key : 'value'
                Value : NPS list of averaged ROIs' NPS lists for respective image.
                Key : 'frequencies'
                Value : list of respective frequencies.
    mean_of_averaged_nps_dict : dict
        For current series folder.
        'values' : NPS list averaged among all images in current series folder.
        'frequencies' : respective frequencies.
    peak_info_dict_ave : dict
        Peak information of NPS list averaged among images in current series folder.
        Keys : 'mean_value' (peak NPS value)
               'mean_freq' (peak NPS frequency)
               'left_dev' (freq distance between peak freq and freq, at which NPS
                           sinks under 60% of peak NPS value when moving to left)
               'right_dev' (freq distance between peak freq and freq, at which NPS
                            sinks under 60% of peak NPS value when moving to right)


    Methods
    -------

    """

    def __init__(self, *, obj_roi, obj_arr, fit_order,
                 crop_perc, useFitting, im_height_in_mm,
                 im_width_in_mm, extensions, trunc_percentage,
                 useCentralCropping, start_freq_range, end_freq_range, step,
                 useTruncation, multipleFiles, pixel_size_in_mm, first_data_set):

        """
        Start initialiazation and sorting of all_roi_dict.

        (See all_roi_dict attribute's description in class GUI).

        :param obj_roi: instance of class GUI
            Used to get access to its attributes:
            - all_roi_dict
            - master
            - image_rect_coord_record
            - array
        :param obj_arr: instance of class StartClass
            Used to get access to its attribute:
            - metadata_subdict
            and method create_base_array()
        :param fit_order: int (1 or 2)
            Order of 2d fit of image for background removal.
        :param crop_perc: float

        :param useFitting: boolean
            Choose method of background removal.
            True: 2d fitting is used (order is defined by attribute
                fit_order);
            False: mean value subtraction is used.
        :param im_height_in_mm: float or string 'undefined'
            Height of current dcm-image in mm.
        :param im_width_in_mm: float or string 'undefined'
            Width of current dcm-image in mm.
        :param extensions: list of strings
            Extensions of files to be searched for in selected order.
            (Probably not needed in this class).
        :param trunc_percentage: float
            Percentage of maximum value, under which NPS is truncated
            at higher frequencies.
            (Probably not needed in this class).
        :param x0: int
            For the mode 'Array_ROIs':
            absolute x coordinate of left upper corner of
            left upper ROI in ROIs array. (???)
        :param x1:
        :param y0:
        :param y1:
        :param useCentralCropping:
        :param start_freq_range: float
            Start frequency of interest. Specified in init_dict.
        :param end_freq_range: float
            End frequency of interest. Specified in init_dict.
        :param step: float
            Interval between frequency samples in freq array.
        :param useTruncation: boolean
            Whether use truncation as explained in description
            of attribute trunc_percentage.
        :param multipleFiles: not needed
        :param pixel_size_in_mm: float
            Default pixel size in mm. used when dealing with
            not-DICOMs.
        :param first_data_set: boolean
            Specify folder structure of dataset.
            Near description is to find in manual.
        """

        print('Constructor of class ProcessROI is being executed')

        # ************Initialization of attributes*********************

        # all xlsx-files
        self.all_xlsx_range = []
        self.all_xlsx = []
        # dictionary for averaged nps and freqs
        self.all_average_nps = {}
        # info about all averaged nps
        self.all_nps_peak_info = {}
        self.all_nps_peak_info_ave = {}
        # paths to all cropped images
        self.all_cropped_im = []
        # paths to all one d nps images
        self.all_1_d_nps = []
        # collect all mean HU for each ROI
        self.image_mean_HU = []
        # collect mean HU for each image
        self.all_mean_HU_dict = {}
        # collect all st dev in dictionary
        self.all_SD_dict = {}
        # dictionary for roi sizes in each image
        self.roi_size_dict = {}
        # dictionary for AUC
        self.auc_dict = {}
        # dictionary for integral of 2d NPS
        self.integral_2d_nps_dict = {}
        # whether central cropping should be applied
        self.useCentralCropping = useCentralCropping
        # whether lower nps should be deleted
        self.useTruncation = useTruncation
        # extensions of image files
        self.extensions = extensions
        # image measurements in mm
        self.im_height_in_mm = im_height_in_mm
        self.im_width_in_mm = im_width_in_mm
        # truncate lower nps
        self.trunc_percentage = trunc_percentage
        # maximal size of the image (height or width)
        self.max_size = max(im_height_in_mm, im_width_in_mm)
        # cropping percentage
        self.crop_perc = crop_perc
        # fitting order for 2d-fit
        self.fit_order = fit_order
        # whether fitting should be applied
        # or background removal should be used
        self.useFitting = useFitting
        # remove raw csv-files
        self.files_to_remove = []
        self.pixel_size_in_mm = pixel_size_in_mm
        self.multipleFiles = multipleFiles
        # declaring attributes, that are specified later
        self.nps = []
        self.new_dict = {}
        # type of data set
        self.first_data_set = first_data_set
        self.object_roi = obj_roi
        self.object_arr = obj_arr
        self.fit_order = fit_order
        self.all_roi = self.object_roi.all_roi_dict
        self.directories_dict = self.object_arr.filedict
        self.metadata_subdict = self.object_arr.metadata_subdict
        # list for paths to Excel-files
        self.xlsx_paths = []

        # numbers of rows and columns of ROI pixel array
        self.px_width = self.object_roi.array.shape[1]
        self.px_height = self.object_roi.array.shape[0]
        # self.arrays_dict = self.object_roi.arrays_dict
        self.all_roi_dict = self.object_roi.all_roi_dict
        self.image_rect_coord = self.object_roi.image_rect_coord
        self.image_rect_coord_record = self.object_roi.image_rect_coord_record

        # ************End of initialization*********************

        # get metadata and prepare for creating summary_info xlsx
        self.headers_list = ['Number', 'Folder', 'Series', 'peak_freq', 'peak_value',
                             'left_dev', 'right_dev', 'area', 'Integral',
                             'ave_m_HU', 'ave_SD']

        self.metadata_headers = [key for key in self.metadata_subdict]
        self.headers_list += self.metadata_headers

        # initialize start row (row of first series in current folder)
        self.start_row = 3

        # create xlsx-file, that will contain summary information
        self.name_workbook_summary = 'Summary_information.xlsx'
        self.workbook_summary = opxl.Workbook()

        # create single worksheet in this workbook
        self.worksheet_summary = self.workbook_summary.active

        # get column letters
        self.col_number = opxl.utils.get_column_letter(1)
        self.worksheet_summary.column_dimensions[self.col_number].width = 6.73
        self.col_folder = opxl.utils.get_column_letter(2)
        self.worksheet_summary.column_dimensions[self.col_folder].width = 30.00
        self.col_series = opxl.utils.get_column_letter(3)
        self.worksheet_summary.column_dimensions[self.col_series].width = 5.18
        self.col_peak_freq = opxl.utils.get_column_letter(4)
        self.worksheet_summary.column_dimensions[self.col_peak_freq].width = 9.09
        self.col_peak_value = opxl.utils.get_column_letter(5)
        self.worksheet_summary.column_dimensions[self.col_peak_value].width = 9.2
        self.col_left_dev = opxl.utils.get_column_letter(6)
        self.col_right_dev = opxl.utils.get_column_letter(7)
        self.col_area = opxl.utils.get_column_letter(8)
        self.col_int_2d_nps = opxl.utils.get_column_letter(9)
        self.col_ave_m_HU = opxl.utils.get_column_letter(10)
        self.worksheet_summary.column_dimensions[self.col_ave_m_HU].width = 10.00
        self.col_ave_SD = opxl.utils.get_column_letter(11)

        self.metadata_columns = [opxl.utils.get_column_letter(i+12) for i
                                 in range(len(self.metadata_headers))]

        for col in self.metadata_columns:
            self.worksheet_summary.column_dimensions[col].width = 33.00

        # name of xlsx-file
        self.name_xlsx = 'NPS_ranged_GUI.xlsx'
        # letters as column names in excel
        self.letters_for_excel = ['A', 'B', 'C', 'D', 'E',
                                  'F', 'G', 'H', 'I', 'J',
                                  'K', 'L', 'M', 'N', 'O',
                                  'P', 'Q', 'R', 'S', 'T',
                                  'U', 'V', 'W', 'X', 'Y',
                                  'Z', 'AA', 'AB', 'AC', 'AD',
                                  'AE', 'AF', 'AG', 'AH', 'AI',
                                  'AJ', 'AK', 'AL']

        # freq range
        self.start_freq = start_freq_range
        self.end_freq = end_freq_range
        self.num_of_steps = int((self.end_freq - self.start_freq) // step + 1)
        self.freq_range = np.linspace(start=self.start_freq,
                                      stop=self.end_freq,
                                      num=self.num_of_steps)
        # sort all_roi_dict
        self.sorted_all_roi_dict = ProcessROI.sort_all_roi_dict(directories_dict=self.directories_dict,
                                                                all_roi_dict=self.all_roi)

        print('Constructor of class ProcessROI is done')

    def execute_calc_nps_sorted(self):

        """
        Main method of the class ProcessROI.

        Calculate NPS by iterating over pixel array of each ROI.
        Create xlsx-files with results.
        :return: nothing
        """

        ave_folder = GUI.create_aux_folder(cur_fold=os.getcwd(), folder_name='Only_averaged_sheets')

        # iterate over keys of the passed dict, i.e. folder paths
        for self.num_folder, self.folder in enumerate(self.sorted_all_roi_dict):
            # log the process
            print('\n\nFolder %s is been processed: %d of %d\n\n' % (os.path.basename(self.folder), self.num_folder + 1,
                                                                     len(self.sorted_all_roi_dict)))

            if self.first_data_set:
                self.folder_part = os.path.basename(self.folder)
            else:
                self.folder_part = ProcessROI.drop_part_of_name(
                    name=os.path.basename(self.folder),
                    pattern_of_dropped_part=r' \- \d+',
                    dropped_from_end=True)

            # create workbook for only averaged sheets
            name_averaged_workbook = ave_folder + '/' + self.folder_part + '.xlsx'
            self.workbook_averaged = xlsx.Workbook(name_averaged_workbook)

            # iterate over series
            for self.num_series, series in enumerate(self.sorted_all_roi_dict[self.folder]):

                global start_time_series
                start_time_series = time.time()

                if self.first_data_set:
                        self.serie_part = serie
                        self.folder_part = os.path.basename(self.folder)
                else:
                    self.serie_part = ProcessROI.drop_part_of_name(
                        name=series,
                        pattern_of_dropped_part=r'\w*\d*_',
                        dropped_from_end=False)[1:]
                    self.folder_part = ProcessROI.drop_part_of_name(
                        name=os.path.basename(self.folder),
                        pattern_of_dropped_part=r' \- \d+',
                        dropped_from_end=True)
                # log the process
                print('\nseries %s: %d of %d\nFolder %d of %d\n' % (series, self.num_series + 1, len(self.sorted_all_roi_dict[self.folder]),
                                                                 self.num_folder + 1,
                                                                 len(self.sorted_all_roi_dict)
                                                                 ))
                # if len(series) > 3:
                #     print('This is not a folder with images. Skip')
                #     continue
                # create folder Results
                # GUI.create_aux_folder(cur_fold=folder, folder_name='Results')



                GUI.create_aux_folder(cur_fold=self.folder, folder_name='Results_%s' % self.folder_part)

                # create worksheet to write averaged data into
                self.worksheet_averaged = self.workbook_averaged.add_worksheet(name=self.serie_part)

                name_for_xlsx = self.folder \
                                + '/Results_%s/' % (self.folder_part) \
                                + self.folder_part +\
                                self.serie_part + '.xlsx'
                # open new workbook in Excel
                self.workbook_series = xlsx.Workbook(name_for_xlsx)
                self.execute_nps_comp(all_roi_dict=self.sorted_all_roi_dict[self.folder][series])
                print('++++++++++\n'
                      'execution time per series: %f seconds\n' % (time.time() - start_time_series))
                # self.execute_nps_comp(all_roi_dict=self.sorted_all_roi_dict[self.folder][series])
                num_remaining_folders = len(self.sorted_all_roi_dict) - self.num_folder
                num_rem_series_in_folder = len(self.sorted_all_roi_dict[self.folder]) - self.num_series - 1
                time_for_one_series = time.time() - start_time_series
                remaining_time = ((num_remaining_folders - 1) * len(self.sorted_all_roi_dict[self.folder]) +
                                  num_rem_series_in_folder) * time_for_one_series
                remaining_hours = remaining_time // 3600
                remaining_minutes = (remaining_time - remaining_hours * 3600) // 60
                remaining_seconds = remaining_time - remaining_minutes * 60
                print(
                    '%d hours, %d minutes, %f seconds remain' % (remaining_hours, remaining_minutes, remaining_seconds))
            self.workbook_averaged.close()
            # increment start row for summary workbook
            self.start_row += self.num_series + 2
        self.workbook_summary.save(self.name_workbook_summary)
        if init_dict['destroy_main_window']:
            self.object_roi.master.destroy()

    def execute_nps_comp(self, all_roi_dict):
        
        """
        Calculate NPS and side variables inside series folder loop.
        
        :param all_roi_dict: dict
            (See attribute sorted_all_roi_dict)
        :return: nothing
        """
        # flush dict of ave nps for the current serie
        self.all_average_nps = {}
        # dictionary to store nps for each image
        self.all_nps_dict = {}

        # flush dictionaries
        self.all_mean_HU_dict = {}
        self.all_SD_dict = {}
        self.integral_2d_nps_dict = {}
        self.auc_dict = {}
        # iterate through all images
        for num_of_image, self.key_image in enumerate(all_roi_dict):

            # initialize list of image ROIs' AUC
            image_auc_list = []
            # initialize list of image ROI's integral of 2d NPS
            image_integral_2d_nps_list = []

            data_from_dicom = self.object_arr.create_base_array(self.key_image)
            metadata_from_dicom = data_from_dicom['whole_dcm']
            try:
                pixel_spacing = [float(i) for i in metadata_from_dicom['0x0028', '0x0030'].value]
            except ValueError:
                pixel_spacing = self.pixel_size_in_mm
                print('There is no property \'Pixel Spacing\'')
            except TypeError:
                pixel_spacing = [0.378, 0.378]
            pixel_array_image = data_from_dicom['base_array']

            # if new series begins
            if num_of_image == 0:
                self.metadata = data_from_dicom['metadata_subdict']

            # build dict of mean HU and SD

            self.build_all_mean_HU_SD_dict(array_to_operate=pixel_array_image,
                                           all_roi_dict=all_roi_dict,
                                           key=self.key_image)
            print('ROIs on image %s are being processed: %d of %d; '
                  'Folder %d of %d; '
                  'series %d of %d ' % (os.path.basename(self.key_image),
                                        num_of_image + 1,
                                        len(all_roi_dict),
                                        self.num_folder + 1,
                                        len(self.sorted_all_roi_dict),
                                        self.num_series + 1,
                                        len(self.sorted_all_roi_dict[self.folder])))
            # counter_roi_inside_image = 0
            # list to store all nps for current image
            self.nps_image = []
            # collect lengths of one d nps of rois in image
            self.lengths = []
            self.image_roi_sizes = []
            # iterate through all rois inside one image
            for num_of_roi, self.item_roi in enumerate(all_roi_dict[self.key_image]):
                subarray = pixel_array_image[self.item_roi[1]:self.item_roi[3], self.item_roi[0]:self.item_roi[2]]
                # print progress
                print('ROI is being processed: %d of %d' % (num_of_roi + 1, len(all_roi_dict[self.key_image])))
                # basename of image without extensions
                self.basename = os.path.basename(self.key_image)[:-4]
                # basename of image with extension
                self.basename_w_ext = os.path.basename(self.key_image)
                # array from the image
                array_to_operate = subarray
                # get shape of the ROI
                shape_of_roi = array_to_operate.shape
                # store the shape in list
                self.image_roi_sizes.append(shape_of_roi)
                # # apply fitting of the image
                # self.create_pol_fit(array_to_operate)
                # create dictionary of nps and respective frequencies (unranged)
                dict = self.compute_nps(array=array_to_operate, pixel_spacing=pixel_spacing)
                AUC = dict['AUC']
                integral_of_2d_NPS = dict['integral_of_2d_NPS']
                # append ROI's AUC und integral of 2d NPS to resp. lists
                image_auc_list.append(dict['AUC'])
                image_integral_2d_nps_list.append(dict['integral_of_2d_NPS'])
                if self.useTruncation:  # setting in init_dict
                    # truncate lower nps and respective frequencies
                    self.new_dict = self.truncate_nps_freq(dict=dict)
                else:
                    # use nps_dict as it is
                    self.new_dict = dict
                # create raw csv-file (with empty rows between data)
                # self.all_xlsx.append(self.create_xlsx_file_nps(dict=self.new_dict, prefix='One_D_NPS_'))
                # create nps range

                # get equations of lines connecting each two points of nps array
                eqs_prop = ProcessROI.nps_equation(self.new_dict['values'])
                # initialize empty list for ranged NPS (with specified s´distance
                # between samples)
                nps_range = []
                # we need this array because freq_range is restricted through
                # size of ROI, frequencies beyond the last available freq
                # should be truncated (dropped)
                new_freq_range = []
                # iterate through frequencies in freq_range
                for item_freq in self.freq_range:
                    try:
                        nps_range.append(self.get_current_nps(freq_array=self.new_dict['frequencies'],
                                                              freq_value=item_freq))
                        new_freq_range.append(item_freq)
                    except ValueError:
                        # if there are no more frequencies available
                        # print('finished')
                        # print('nps range: ', len(nps_range))
                        break
                # nps_range = list(map(fut.partial(self.get_current_nps, freq_array=self.new_dict['frequencies'],
                #                                                            eqs=eqs_prop), self.freq_range))
                range_dict = {'values': nps_range,
                              'frequencies': new_freq_range,
                              'AUC': AUC,
                              'integral_of_2d_NPS': integral_of_2d_NPS}

                # store ranged nps and resp. freq in a list
                self.nps_image.append(range_dict)
                assert len(nps_range) == len(new_freq_range)
                self.lengths.append(len(range_dict['values']))
                # print('continued:  ', self.lengths)

            # update dict for AUC and integral of 2d NPS
            self.auc_dict.update({self.key_image: image_auc_list})
            self.integral_2d_nps_dict.update({self.key_image: image_integral_2d_nps_list})

            # average stored nps
            averaged_dict = self.average_roi_nps(list_of_dict=self.nps_image)
            self.all_average_nps.update({self.key_image: averaged_dict})
            self.roi_size_dict.update({self.key_image: self.image_roi_sizes})

            # recognize all peaks in nps-array
            peaks = ProcessROI.collect_all_max_peaks_nps(averaged_dict)
            # handle peak info
            peak_info_dict = self.handle_peak_info(peak_dict=peaks,
                                                   all_val_arr=averaged_dict['values'],
                                                   all_freq_arr=averaged_dict['frequencies'])
            self.all_nps_peak_info.update({self.key_image: peak_info_dict})

            self.all_nps_dict.update({self.key_image: self.nps_image})
        # create mean HU and SD info dictionaries
        # self.build_all_mean_HU_SD_dict(all_roi_dict=all_roi_dict)
        # self.build_all_sd_dict(all_roi_dict=all_roi_dict,
        #                        all_mean_dict=self.all_mean_HU_dict)
        # print('SD:  ', self.all_SD_dict)
        # print('mean HU:  ', self.all_mean_HU_dict)
        # calculate mean of averaged nps
        self.mean_of_averaged_nps_dict = ProcessROI.mean_of_ave_nps(all_average_nps=self.all_average_nps)
        # recognize all peaks in nps-array
        peaks_ave = ProcessROI.collect_all_max_peaks_nps(self.mean_of_averaged_nps_dict)
        # handle peak info
        self.peak_info_dict_ave = self.handle_peak_info(peak_dict=peaks_ave,
                                                        all_val_arr=self.mean_of_averaged_nps_dict['values'],
                                                        all_freq_arr=self.mean_of_averaged_nps_dict['frequencies'])
        # self.all_nps_peak_info_ave.update({self.key_image: self.peak_info_dict_ave})
        # calculate SD of mean HU
        self.sd_of_mean_HU_dict = ProcessROI.sd_of_dictionary(dict=self.all_mean_HU_dict)
        # calculate SD of SD
        self.sd_of_sd_dict = ProcessROI.sd_of_dictionary(dict=self.all_SD_dict)
        # get total mean values for mean_HU and SD
        self.total_mean_HU = ProcessROI.mean_of_mean(all_values_dict=self.all_mean_HU_dict)
        self.total_mean_sd = ProcessROI.mean_of_mean(all_values_dict=self.all_SD_dict)
        # create workbook for displaying results
        # self.workbook_series = xlsx.Workbook(self.name_xlsx)
        self.create_xlsx_file_nps(all_nps_dict=self.all_nps_dict)
        self.workbook_series.close()
        pass
        return

    @staticmethod
    def sort_all_roi_dict(directories_dict, all_roi_dict):
        """
        Sort passed all_roi_dict to reproduce
        directory structure of dcm-images data set
        (see description below).

        :param directories_dict: dict of dicts of lists of strings
            (See description of attribute filedict of class StartClass)
        :param all_roi_dict: dict of lists of tuples
            (See decription of attribute all_roi_dict of class GUI)
        :return: sorted all_roi_dict
            Keys : paths to study folder
            Values : dicts
                Keys : paths to series folders
                Values : dict
                    Keys : paths to image file
                    Values : lists of tuples containing coordinates of ROIs:
                        - x coordinate of upper left corner
                        - y coordinate of upper left corner
                        - x coordinate of lower right corner
                        - y coordinate of lower right corner

        """

        print('sort_all_roi_dict is being executed')
        # empty dict for sorted ROIs
        sorted_all_roi_dict = {}
        total_files = len(all_roi_dict.keys())
        # iterate over keys of all_roi_dict, i.e. file names
        for numf, file_name_prim in enumerate(natsorted(all_roi_dict.keys(), key=lambda f: f.split('_')[-1])):
            local_start_time = time.time()
            print('Progress: file %d of %d' % ((numf + 1), total_files))
            # iterate over the keys of directories_dict
            for classdirname in natsorted(directories_dict.keys(), key=lambda f: f.split('_')[-1]):
                # update dict
                # sorted_all_roi_dict.update({classdirname: {}})
                # iterate over keys of subdict
                subdict = directories_dict[classdirname]
                for serie_name in natsorted(subdict.keys(), key=lambda f: f.split('_')[-1]):
                    # update dict
                    # sorted_all_roi_dict[classdirname].update({serie_name: {}})
                    # iterate over the files in file list
                    for file_name_second in subdict[serie_name]:
                        # compare key of all_roi_dict and filename
                        if file_name_prim in file_name_second:
                            try:
                                temp_dict_prim = sorted_all_roi_dict[classdirname]
                            except KeyError:
                                sorted_all_roi_dict.update({classdirname: {}})
                                temp_dict_prim = sorted_all_roi_dict[classdirname]

                            try:
                                temp_dict = temp_dict_prim[serie_name]
                            except KeyError:
                                temp_dict_prim.update({serie_name: {}})
                                temp_dict = temp_dict_prim[serie_name]

                            temp_dict.update({file_name_prim: all_roi_dict[file_name_prim]})
                            sorted_all_roi_dict[classdirname].update({serie_name: temp_dict})
            duration_for_loop = time.time() - local_start_time
        print('sort_all_roi_dict is done')

        return sorted_all_roi_dict

    def build_all_mean_HU_SD_dict(self, all_roi_dict, array_to_operate, key):

        """
        Calculate mean HU and standard deviation for each ROI
        on the current image and update respective dictionaries
        all_mean_HU_dict and all_SD_dict (See description in class' docs)

        :param all_roi_dict: dict
            (See description of attribute sorted_all_roi_dict).
        :param array_to_operate: ndarray (2d)
            Current ROI's pixel array.
        :param key: string
            Path to current dcm-image file.
        :return: nothing
        """

        # all mean HU for the current image
        roi_image_mean_HU = []
        # all mean sd for current image
        image_sd = []
        # iterate through all rois in image
        for coord in all_roi_dict[key]:
            # get pixel array of current roi
            roi_array = array_to_operate[coord[1]:coord[3], coord[0]:coord[2]]

            # calculate mean HU
            mean_HU = np.mean(roi_array)
            roi_image_mean_HU.append(mean_HU)

            # calculate SD
            # build homogen mean matrix
            mean_matrix = np.ones(shape=np.array(roi_array).shape) * mean_HU
            # calculate difference between roi image and mean image
            diff_matrix = roi_array - mean_matrix
            # flatten diff matrix to access all its elements easier
            diff_flattened = diff_matrix.ravel()
            # calculate SD for current ROI
            sd_roi = np.sqrt(np.mean([i ** 2 for i in diff_flattened]))
            image_sd.append(sd_roi)
        self.all_mean_HU_dict.update({key: roi_image_mean_HU})
        self.all_SD_dict.update({key: image_sd})

        pass

    @staticmethod
    def mean_of_ave_nps(all_average_nps):
        """
        Averaging of passed nps values dict
        among all images in current series folder.
        :param all_average_nps: dict of lists of dicts
            Keys : paths to images in current series folder.
            Values : list range_dict attribute for each ROI.
                (See attribute range_dict)
        :return: dict
            (Key :) 'values' : (Value :) nps list averaged
        """
        averaged_nps_as_list = []
        for key in all_average_nps:
            averaged_nps_as_list.append(all_average_nps[key]['values'])
            frequens = all_average_nps[key]['frequencies']
        mean_of_averaged_nps = np.mean(np.array(averaged_nps_as_list), axis=0)
        mean_of_averaged_nps_dict = {'values': mean_of_averaged_nps,
                                     'frequencies': frequens}
        return mean_of_averaged_nps_dict

    @staticmethod
    def sd_of_dictionary(dict):
        """
        Calculate standard deviation of list values of given dict.
        
        :param dict: dict
            Keys : whatever keys;
            Values : lists of numeric values;
        :return: dict
            Keys : the same keys as of dict argument;
            Values : sd value of respective list.
        """
        
        # build array from dictionary
        sd_dict = {}
        for key in dict:
            one_d_array = dict[key]
            mean_value = np.mean(one_d_array)
            squared_diff_list = []
            for item in one_d_array:
                squared_diff = (item - mean_value) ** 2
                squared_diff_list.append(squared_diff)
            sd = np.sqrt(np.mean(squared_diff_list))
            sd_dict.update({key: sd})

        return sd_dict

    def average_roi_nps(self, list_of_dict):
        """
        Build dictionary of nps lists averaged among ROIs in each image.

        :param list_of_dict: list_of_dict
            (See attribute nps_image)
        :return: dict of dicts
            Keys : absolute paths to images in current series folder.
            Values : dict
                Key : 'value'
                Value : NPS list of averaged ROIs' NPS lists for respective image.
                Key : 'frequencies'
                Value : list of respective frequencies.
        """

        # initialize lists for nps and resp freqs
        values_to_average = []
        resp_freq_to_average = []
        # get max length of roi nps and its index
        max_length = max(self.lengths)
        max_length_idx = np.argmax(self.lengths)
        # iterate through all rois nps in image
        for roi_item_dict in list_of_dict:
            values = roi_item_dict['values']
            frequencies = roi_item_dict['frequencies']

            # fill lacking items with zeros

            # length difference with the longest nps
            len_diff = max_length - len(values)
            # convert values und freqs to python lists
            values = list(values)
            frequencies = list(frequencies)
            values += [0] * len_diff
            frequencies += [0] * len_diff
            # store transformed nps values and freqs in lists
            values_to_average.append(values)
            resp_freq_to_average.append(frequencies)
        # get mean array of value arrays
        averaged_nps = np.mean(values_to_average, axis=0)
        # max_length_idx tells us, which ROI has the largest NPS range
        # then the frequencies array of this ROI is retrieved from nps_image
        averaged_freqs = list_of_dict[max_length_idx]['frequencies']
        # store averaged value and frequencies in one dictionary
        averaged_dict = {'values': averaged_nps,
                         'frequencies': averaged_freqs}
        return averaged_dict

    # method is used to get total average values from all_mean_HU_dict and all_sd_dict
    @staticmethod
    def mean_of_mean(all_values_dict):

        """
        Build mean value of numerical values of given dictionary.
        :param all_values_dict: dict
            Keys : Any
            Values : single numerical values
        :return: mean value of the values
        """

        mean_of_mean_list = []
        for key in all_values_dict:
            mean_of_mean_list.append(np.mean(all_values_dict[key]))
        mean_of_mean_value = np.mean(mean_of_mean_list)
        return mean_of_mean_value

    def compute_nps(self, array, pixel_spacing):
        
        """
        Compute 2d and 1d NPS of given pixel array.
        
        :param array: ndarray (2d)
            Pixel array of current ROI. 
        :param pixel_spacing: tuple of two floats
            Pixel spacing of dcm-image in y and 
            x direction.
        :return: dict
            Keys : 'values' - 1d NPS of ROI (not interpolated),
                   'frequencies' - respective frequencies,
                   'AUC' - area under 1d NPS profile,
                   'integral_of_2d_nps' - as in the name.       
        """
        
        # if image measurements in mm are undefined
        # the default image sizing is applied
        if self.im_width_in_mm == 'undefined':
            self.im_width_in_mm = self.px_width * self.pixel_size_in_mm
        if self.im_height_in_mm == 'undefined':
            self.im_height_in_mm = self.px_height * self.pixel_size_in_mm
        # mean pixel value of whole image
        mean_value = np.mean(array)
        # transform list into numpy array
        # to access the array's properties
        np_arr = np.array(array)
        # get ROI size
        roi_width = np_arr.shape[0]
        roi_height = np_arr.shape[1]
        # maximal size of the array (height or width)
        max_size = max(np_arr.shape)
        # building mean value array (background)
        mean_arr = mean_value * np.ones(np_arr.shape)
        # if 2d fitting should be used
        if self.useFitting:
            # self.polyfit is the 2d-fit of the image
            detrended_arr = array - self.pol_fit
        else:
            detrended_arr = array - mean_arr
        # create file of detrended image
        # StartClass.create_image_from_2d_array(arr_2d=detrended_arr,
        #                             filename='09.Detrended_images/Detrended_image__' +
        #                                      self.basename + '__.png')
        # apply FFT to detrended image
        DFT_list = np.fft.fftshift(np.fft.fft2(detrended_arr))
        # calculate 2d-NPS
        # nps = 1 / self.px_width / self.px_height * np.abs(DFT_list)**2
        # nps = (self.pixel_size_in_mm ** 2) / self.px_width / self.px_height * np.abs(DFT_list) ** 2
        nps = 1 / roi_height**2 / roi_width**2 * np.abs(DFT_list) ** 2
        integral_of_2d_NPS = np.sum(nps)

        # create file of 2d-NPS-image
        StartClass.create_image_from_2d_array(arr_2d=nps,
                                        filename='01.2d_NPS_images/NPS_2D__' + 
                                                 self.basename + '__.jpg')
        # building 1d-NPS from 2d_NPS using radial average
        nps_1d = ProcessROI.radial_mean(nps)
        AUC = np.sum(nps_1d)
        # self.nps_norm = self.norm_array(arr_to_normalize=nps_1d,
        #                                 all_val_array=nps_1d)
        # freqs = np.fft.fftfreq(max_size, self.im_width_in_mm/10/self.px_width)[:max_size // 2]
        # calculate respective frequencies (line pairs per cm)
        # freqs = np.fft.fftfreq(max_size, self.im_width_in_mm/10/self.px_width)[:max_size // 2]
        freqs = np.fft.fftfreq(max_size, pixel_spacing[0] / 10)[:max_size // 2]
        # dictionary with all NPS- and freq-values, that will be
        # truncated afterwards
        nps_dict = {'values': nps_1d,
                    'frequencies': freqs,
                    'integral_of_2d_NPS': integral_of_2d_NPS,
                    'AUC': AUC}
        return nps_dict

    @staticmethod
    def drop_part_of_name(name, pattern_of_dropped_part, dropped_from_end):
        
        """
        Recognize part of name and drop it.
        
        :param name: string
            String to be truncated.
        :param pattern_of_dropped_part: raw string
            RegEx-pattern of part to be dropped.
        :param dropped_from_end: boolean
            True: if dropped part is at the left end of string.
            False: if dropped part is at the right end of string.
        :return: string
            Truncated string.
        """
        
        # DEBUG_dropped_part = re.findall(pattern=pattern_of_dropped_part, string=name)
        dropped_part = re.findall(pattern=pattern_of_dropped_part, string=name)[0]
        if dropped_from_end:
            used_part = name[:-(len(dropped_part))]
        else:
            used_part = name[(len(dropped_part) - 1):]
        return used_part

    def create_xlsx_file_nps(self, all_nps_dict):

        """
        Create several xlsx-files with results:

            - <foldername_seriesname>.xlsx (for each series folder;
                                            with info on all ROIs' NPS-values, ended with 'averaged'-worksheet
                                            saved in folder 'Results' inside each study-folder)
            - <foldername>.xlsx (for each study;
                                 with all collected averaged-worksheets from previous files
                                 saved in executive file's dir in dir 'Only_averaged_worksheets')
            - Summary_information.xlsx (with summarized information from foldername.xlsx-files
                                        saved in executive file's dir)
        
        :param all_nps_dict: dict
            Dictionary containing NPS info for each ROI in each image. Has following
            structure:

            {'path_to_image_0': [roi_0 = {'values': [],
                                        'frequencies' [],
                                        'AUC': ...,
                                        'integral_of_2d_NPS':...},
                                roi_1 = {'values': [],
                                        'frequencies' [],
                                        'AUC': ...,
                                        'integral_of_2d_NPS':...},
                            ...],
            'path_to_image_1': ...}
         
        :return: nothing
        """

        # iterate through all images
        for num_of_image, image_key in enumerate(all_nps_dict):
            # print progress
            print('Writing worksheets in xlsx-file: %d of %d' % (num_of_image + 1, len(all_nps_dict)))
            # counter for roi in image multiplied by 2
            counter_roi = 0
            worksheet = self.workbook_series.add_worksheet(os.path.basename(image_key)[:-4])
            for item_nps_dict in all_nps_dict[image_key]:
                val_arr = item_nps_dict['values']
                freq_arr = item_nps_dict['frequencies']

                # initialization of cells in worksheet
                row = 2
                col = counter_roi

                # headers of the table
                worksheet.write(1, counter_roi, 'Lp')
                worksheet.write(1, counter_roi + 1, 'NPS')
                worksheet.write(0, counter_roi, 'ROI_%d' % (counter_roi // 2 + 1))
                worksheet.write(0, counter_roi + 1, '%dx%d px' % (self.roi_size_dict[image_key][counter_roi // 2][0],
                                                                    self.roi_size_dict[image_key][counter_roi//2][1]))

                for frequency, value_nps in zip(freq_arr, val_arr):
                    worksheet.write(row, col, frequency)
                    worksheet.write(row, col + 1, value_nps)
                    row += 1  # next row

                counter_roi += 2
            # retrieve mean of AUC and integral of 2d NPS for the current image
            AUC = np.mean(self.auc_dict[image_key])
            integral_of_2d_NPS = np.mean(self.integral_2d_nps_dict[image_key])

            row_ave = 2
            col_ave = counter_roi + 1
            # averaged nps data
            worksheet.write(row_ave - 2, col_ave, 'averaged')
            worksheet.write(row_ave - 1, col_ave, 'Lp')
            worksheet.write(row_ave - 1, col_ave + 1, 'NPS')
            for frequency, value_nps in zip(self.all_average_nps[image_key]['frequencies'],
                                            self.all_average_nps[image_key]['values']):
                worksheet.write(row_ave, col_ave, frequency)
                worksheet.write(row_ave, col_ave + 1, value_nps)
                row_ave += 1  # next row

            # create a new Chart object
            chart = self.workbook_series.add_chart({'type': 'line'})
            # configure the chart
            chart.add_series({'values': '=%s!$%s$3:$%s$%d' % (os.path.basename(image_key)[:-4],
                                                              self.letters_for_excel[counter_roi + 2],
                                                              self.letters_for_excel[counter_roi + 2],
                                                              len(self.all_average_nps[image_key]['frequencies']) + 2),
                              'categories': '%s!$%s$3:$%s$%d' % (os.path.basename(image_key)[:-4],
                                                                 self.letters_for_excel[counter_roi + 1],
                                                                 self.letters_for_excel[counter_roi + 1],
                                                                 len(self.all_average_nps[image_key]['frequencies']) + 2),
                              'name': os.path.basename(image_key),
                              'legend': False,
                              'trendline': {'type': 'polynomial',
                                            'order': 3,
                                            'line': {
                                                'color': 'red',
                                                'width': 1,
                                                'dash_type': 'long_dash',
                                            },
                                            'display_equation': True,
                                            }})
            chart.set_x_axis({'name': 'Line pairs per cm'})
            chart.set_y_axis({'name': 'NPS_1D_averaged'})
            # Insert the chart into the worksheet.
            worksheet.insert_chart('%s1' % self.letters_for_excel[counter_roi + 3], chart)

            # additional information about size of cropped image
            # and characteristics of nps curve

            worksheet.write(19, counter_roi + 4, 'max_peak_nps')
            worksheet.write(20, counter_roi + 4, 'max_peak_freq')
            worksheet.write(19, counter_roi + 5, self.all_nps_peak_info[image_key]['mean_value'])
            worksheet.write(20, counter_roi + 5, self.all_nps_peak_info[image_key]['mean_freq'])
            worksheet.write(21, counter_roi + 4, 'left_dev')
            worksheet.write(22, counter_roi + 4, 'right_dev')
            worksheet.write(21, counter_roi + 5, self.all_nps_peak_info[image_key]['left_dev'])
            worksheet.write(22, counter_roi + 5, self.all_nps_peak_info[image_key]['right_dev'])

            # area under NPS-curve
            worksheet.write(24, counter_roi + 4, 'area')
            worksheet.write(24, counter_roi + 5, AUC)

            # integral of 2d NPS
            worksheet.write(25, counter_roi + 4, 'Integral_2d_NPS')
            worksheet.write(25, counter_roi + 5, integral_of_2d_NPS)

            # fitting info
            if self.useFitting:
                worksheet.write(17, counter_roi + 4, 'Fitting')
                worksheet.write(17, counter_roi + 5, self.fit_order)
            else:
                worksheet.write(17, counter_roi + 4, 'BG_remove')
            # make column wider
            worksheet.set_column(first_col=counter_roi + 4, last_col=counter_roi + 4, width=20)

            # write info about mean HU and standard deviation
            row_mean_HU_SD_info = 19
            col_mean_HU_SD_info_header = counter_roi + 7
            col_mean_HU = col_mean_HU_SD_info_header + 1
            col_SD = col_mean_HU + 1
            col_x_coord = col_SD + 1
            col_y_coord = col_x_coord + 1
            for mean_HU, SD, coord in zip(self.all_mean_HU_dict[image_key],
                                   self.all_SD_dict[image_key],
                                          self.image_rect_coord_record):
                worksheet.write(row_mean_HU_SD_info, col_mean_HU_SD_info_header,
                                'ROI_%d' % (row_mean_HU_SD_info - 18))
                worksheet.write(row_mean_HU_SD_info, col_mean_HU, mean_HU)
                worksheet.write(row_mean_HU_SD_info, col_SD, SD)
                worksheet.write(row_mean_HU_SD_info, col_x_coord, coord[0])
                worksheet.write(row_mean_HU_SD_info, col_y_coord, coord[1])
                row_mean_HU_SD_info += 1
            worksheet.write(18, counter_roi + 8, 'Mean_HU')
            worksheet.write(18, counter_roi + 9, 'SD')
            worksheet.write(18, counter_roi + 10, 'x_coord')
            worksheet.write(18, counter_roi + 11, 'y_coord')

            worksheet.write(27, counter_roi + 4, 'averaged_Mean_HU')
            worksheet.write(28, counter_roi + 4, 'averaged_SD')
            worksheet.write(29, counter_roi + 4, 'SD_of_Mean_HU')
            worksheet.write(30, counter_roi + 4, 'SD_of_SD')

            worksheet.write(row_mean_HU_SD_info + 1, col_mean_HU_SD_info_header,
                            'averaged')
            worksheet.write(row_mean_HU_SD_info + 2, col_mean_HU_SD_info_header,
                            'SD')
            worksheet.write(row_mean_HU_SD_info + 1, col_mean_HU,
                            np.mean(self.all_mean_HU_dict[image_key]))
            worksheet.write(row_mean_HU_SD_info + 1, col_SD,
                            np.mean(self.all_SD_dict[image_key]))

            worksheet.write(27, counter_roi + 5, np.mean(self.all_mean_HU_dict[image_key]))
            worksheet.write(28, counter_roi + 5, np.mean(self.all_SD_dict[image_key]))
            worksheet.write(29, counter_roi + 5, np.mean(self.sd_of_mean_HU_dict[image_key]))
            worksheet.write(30, counter_roi + 5, np.mean(self.sd_of_sd_dict[image_key]))

            # write sd of mean HU and SD
            worksheet.write(row_mean_HU_SD_info + 2, col_mean_HU,
                            np.mean(self.sd_of_mean_HU_dict[image_key]))
            worksheet.write(row_mean_HU_SD_info + 2, col_SD,
                            np.mean(self.sd_of_sd_dict[image_key]))
        worksheet_ave = self.workbook_series.add_worksheet('averaged')
        val_arr = self.mean_of_averaged_nps_dict['values']
        freq_arr = self.mean_of_averaged_nps_dict['frequencies']

        # initialization of cells in worksheet
        row = 2
        col = 0

        # headers of the table
        worksheet_ave.write(0, 0, 'Total average')
        worksheet_ave.write(1, 0, 'Lp')
        worksheet_ave.write(1, 0 + 1, 'NPS')

        self.worksheet_averaged.write(0, 0, 'Total average')
        self.worksheet_averaged.write(1, 0, 'Lp')
        self.worksheet_averaged.write(1, 0 + 1, 'NPS')
        # worksheet.write(0, 1, 'ROI_%d' % (1 // 2 + 1))
        # worksheet.write(0, 1 + 1,
        #                 '%dx%d px' % (self.roi_size_dict[image_key][1 // 2][0],
        #                               self.roi_size_dict[image_key][1 // 2][1]))

        for frequency, value_nps in zip(freq_arr, val_arr):
            worksheet_ave.write(row, col, frequency)
            worksheet_ave.write(row, col + 1, value_nps)

            self.worksheet_averaged.write(row, col, frequency)
            self.worksheet_averaged.write(row, col + 1, value_nps)
            row += 1  # next row


        # additional information about size of cropped image
        # and characteristics of nps curve

        worksheet_ave.write(19 - 4, 1 + 3, 'max_peak_nps')
        worksheet_ave.write(20 - 4, 1 + 3, 'max_peak_freq')
        worksheet_ave.write(19 - 4, 1 + 4, self.peak_info_dict_ave['mean_value'])
        worksheet_ave.write(20 - 4, 1 + 4, self.peak_info_dict_ave['mean_freq'])
        worksheet_ave.write(21 - 4, 1 + 3, 'left_dev')
        worksheet_ave.write(22 - 4, 1 + 3, 'right_dev')
        worksheet_ave.write(21 - 4, 1 + 4, self.peak_info_dict_ave['left_dev'])
        worksheet_ave.write(22 - 4, 1 + 4, self.peak_info_dict_ave['right_dev'])

        self.worksheet_averaged.write(19 - 4, 1 + 3, 'max_peak_nps')
        self.worksheet_averaged.write(20 - 4, 1 + 3, 'max_peak_freq')
        self.worksheet_averaged.write(19 - 4, 1 + 4, self.peak_info_dict_ave['mean_value'])
        self.worksheet_averaged.write(20 - 4, 1 + 4, self.peak_info_dict_ave['mean_freq'])
        self.worksheet_averaged.write(21 - 4, 1 + 3, 'left_dev')
        self.worksheet_averaged.write(22 - 4, 1 + 3, 'right_dev')
        self.worksheet_averaged.write(21 - 4, 1 + 4, self.peak_info_dict_ave['left_dev'])
        self.worksheet_averaged.write(22 - 4, 1 + 4, self.peak_info_dict_ave['right_dev'])

        # writing info averaged Mean_HU, averaged SD, and area
        worksheet_ave.write(24 - 4, 1 + 3, 'Int of 2d-NPS')
        worksheet_ave.write(26 - 4, 1 + 3, 'averaged Mean_HU')
        worksheet_ave.write(27 - 4, 1 + 3, 'averaged SD')

        worksheet_ave.write(24 - 4, 1 + 4, np.mean(
            [self.integral_2d_nps_dict[key] for key in self.integral_2d_nps_dict]
        ))
        worksheet_ave.write(26 - 4, 1 + 4, self.total_mean_HU)
        worksheet_ave.write(27 - 4, 1 + 4, self.total_mean_sd)

        self.worksheet_averaged.write(24 - 4, 1 + 3, 'Int of 2d-NPS')
        self.worksheet_averaged.write(26 - 4, 1 + 3, 'averaged Mean_HU')
        self.worksheet_averaged.write(27 - 4, 1 + 3, 'averaged SD')

        self.worksheet_averaged.write(24 - 4, 1 + 4, np.mean(
            [self.integral_2d_nps_dict[key] for key in self.integral_2d_nps_dict]
        ))
        self.worksheet_averaged.write(26 - 4, 1 + 4, self.total_mean_HU)
        self.worksheet_averaged.write(27 - 4, 1 + 4, self.total_mean_sd)


        # make column wider
        worksheet_ave.set_column(first_col=4, last_col=4, width=20)
        self.worksheet_averaged.set_column(first_col=4, last_col=4, width=20)


        # info about averaged mean_HU and SD
        worksheet_ave.write(19 - 4, 1 + 7, 'mean_HU')
        worksheet_ave.write(19 - 4, 1 + 8, 'SD')
        worksheet_ave.write(20 - 4, 1 + 6, 'averaged')
        worksheet_ave.write(20 - 4, 1 + 7, self.total_mean_HU)
        worksheet_ave.write(20 - 4, 1 + 8, self.total_mean_sd)

        self.worksheet_averaged.write(19 - 4, 1 + 7, 'mean_HU')
        self.worksheet_averaged.write(19 - 4, 1 + 8, 'SD')
        self.worksheet_averaged.write(20 - 4, 1 + 6, 'averaged')
        self.worksheet_averaged.write(20 - 4, 1 + 7, self.total_mean_HU)
        self.worksheet_averaged.write(20 - 4, 1 + 8, self.total_mean_sd)

        # create a new Chart object
        chart_ave = self.workbook_series.add_chart({'type': 'line'})
        chart_averaged = self.workbook_averaged.add_chart({'type': 'line'})
        # configure the chart
        chart_ave.add_series({'values': '=%s!$%s$3:$%s$%d' % ('averaged',
                                                          'B',
                                                          'B',
                                                          len(self.mean_of_averaged_nps_dict['frequencies']) + 2),
                          'categories': '%s!$%s$3:$%s$%d' % ('averaged',
                                                             'A',
                                                             'A',
                                                             len(self.mean_of_averaged_nps_dict['frequencies']) + 2),
                          'name': 'Total Average',
                          'legend': False,
                          'trendline': {'type': 'polynomial',
                                        'order': 3,
                                        'line': {
                                            'color': 'red',
                                            'width': 1,
                                            'dash_type': 'long_dash',
                                        },
                                        'display_equation': False,
                                        }})

        chart_averaged.add_series({'values': '=%s!$%s$3:$%s$%d' % (self.serie_part,
                                                                   'B',
                                                                   'B',
                                                                   len(self.mean_of_averaged_nps_dict['frequencies']) +
                                                                   2),
                                   'categories': '%s!$%s$3:$%s$%d' % (self.serie_part,
                                                                      'A',
                                                                      'A',
                                                                      len(self.mean_of_averaged_nps_dict[
                                                                              'frequencies']) + 2),
                                   'name': 'Total Average',
                                   'legend': False,
                                   })

        chart_ave.set_x_axis({'name': 'Line pairs per cm'})
        chart_ave.set_y_axis({'name': 'NPS_1D_averaged'})
        # Insert the chart into the worksheet.
        worksheet_ave.insert_chart('C1', chart_ave)

        chart_averaged.set_x_axis({'name': 'Line pairs per cm'})
        chart_averaged.set_y_axis({'name': 'NPS_1D_averaged'})
        # Insert the chart into the worksheet.
        self.worksheet_averaged.insert_chart('C1', chart_averaged)

        # create summary info xlsx

        # write headers to worksheet
        for num_header, header_name in enumerate(self.headers_list):
            self.worksheet_summary['%s1' % opxl.utils.get_column_letter(num_header + 1)] = header_name

        # row to write
        row_to_write = self.start_row + self.num_series
        # get worksheet's name
        name_of_folder = self.folder_part
        name_of_series = self.serie_part
        # write information from worksheet
        self.worksheet_summary['%s%d' % (self.col_number, row_to_write)] = self.num_folder + 1
        self.worksheet_summary['%s%d' % (self.col_folder, row_to_write)] = name_of_folder
        self.worksheet_summary['%s%d' % (self.col_series, row_to_write)] = name_of_series
        self.worksheet_summary['%s%d' % (self.col_peak_freq, row_to_write)] = self.peak_info_dict_ave['mean_freq']
        self.worksheet_summary['%s%d' % (self.col_peak_value, row_to_write)] = self.peak_info_dict_ave['mean_value']
        self.worksheet_summary['%s%d' % (self.col_left_dev, row_to_write)] = self.peak_info_dict_ave['left_dev']
        self.worksheet_summary['%s%d' % (self.col_right_dev, row_to_write)] = self.peak_info_dict_ave['right_dev']
        self.worksheet_summary['%s%d' % (self.col_area, row_to_write)] = np.mean(
            [self.auc_dict[key] for key in self.auc_dict]
        )
        self.worksheet_summary['%s%d' % (self.col_int_2d_nps, row_to_write)] = np.mean(
            [self.integral_2d_nps_dict[key] for key in self.integral_2d_nps_dict]
        )
        self.worksheet_summary['%s%d' % (self.col_ave_m_HU, row_to_write)] = self.total_mean_HU
        self.worksheet_summary['%s%d' % (self.col_ave_SD, row_to_write)] = self.total_mean_sd

        for num_metadata, (metadata_tag, col_metadata) in enumerate(
            zip(self.metadata_headers, self.metadata_columns)
        ):
            self.worksheet_summary['%s%d' % (col_metadata, row_to_write)] = self.metadata[metadata_tag]

    @staticmethod
    def radial_mean(array):
        
        """
        Build radial mean of 2d-array. In our case: 2d-NPS.
        
        
        :param array: ndarray (2d)
            Two-dimensional NPS of current ROI.
        :return: ndarray (1d)
            Radial mean of the 2d-NPS.
        """
        
        image = array
        image_height = array.shape[0]
        image_width = array.shape[1]
        center_x = image_width // 2
        center_y = image_height // 2
        max_size = max(image_height, image_width)
        # create array of radii
        x, y = np.meshgrid(np.arange(image.shape[1]), np.arange(image.shape[0]))
        R = np.sqrt((x - center_x) ** 2 + (y - center_y) ** 2)

        # calculate the mean
        f = lambda r: image[(R >= r - .5) & (R < r + .5)].mean()
        # r = np.linspace(1, 302, num=302)
        r = np.linspace(0, max_size//2, num=max_size//2+1)
        mean = np.vectorize(f)(r)
        mean = [array[center_y][center_x]] + mean
        return mean

    def polyfit2d(self, x, y, z):
        """2d-fitting of 2d-array. Used for background extraction"""
        size_x = np.array(x).shape[0]
        size_y = np.array(y).shape[0]
        if size_x > size_y:
            y = np.concatenate((y, [y[-1]]*(size_x - size_y)))
        else:
            x = np.concatenate((x, [x[-1]] * (size_y - size_x)))
        order = self.fit_order
        ncols = (order + 1) ** 2
        G = np.zeros((x.size, ncols))
        ij = itertools.product(range(order + 1), range(order + 1))
        for k, (i, j) in enumerate(ij):
            G[:, k] = x ** i * y ** j
        try:
            m, _, _, _ = np.linalg.lstsq(G, z, rcond=None)
        except:
            print('There is a problem in file (fitting):   ', file)
        return m

    @staticmethod
    def polyval2d(self, x, y, m):
        """Auxiliar function for 2d-fitting"""
        size_x = np.array(x).shape[0]
        size_y = np.array(y).shape[0]
        if size_x > size_y:
            y = np.concatenate((y, [y[-1]]*(size_x - size_y)))
        else:
            x = np.concatenate((x, [x[-1]] * (size_y - size_x)))
        order = int(np.sqrt(len(m))) - 1
        ij = itertools.product(range(order + 1), range(order + 1))
        z = np.zeros_like(x)
        for a, (i, j) in zip(m, ij):
            z += a * x ** i * y ** j
        return z

    def prepare_f_1(self, xy, a, b, c, d):
        """Auxiliar function for 2d-fitting"""
        i = xy // self.image_width_1  # reconstruct y coordinates
        j = xy % self.image_width_1  # reconstruct x coordinates
        out = i * a + j * b + i * j * c + d
        return out

    @staticmethod
    def nps_equation(nps_array):
        """Input: not ranged nps array (with larger distance between
        samples). Return list of tuples, containing slope and bias of
        line connecting each two points of nps array"""
        line_equations = []
        for s in range(len(nps_array) - 1):
            line_equations.append(ProcessROI.determine_line_equation(point_index_1=s,
                                                                     point_index_2=s+1,
                                                                     point_val_1=nps_array[s],
                                                                     point_val_2=nps_array[s+1]))
        return line_equations

    def get_current_nps(self, freq_value, freq_array):

        """
        Get NPS value respective to frequency in freq range.

        :param freq_value: float
            Current frequency value of freq_range.
        :param freq_array: list of floats
            Not interpolated list of NPS frequencies.
        :return: float
            Interpolated value of NPS respective to given freq_value.
        """

        """
        using equations of lines between points of unranged NPS array"""
        # all frequencies, that less than freq_value
        less_values = [i for i in freq_array if i <= freq_value]
        # all frequencies, that greater than freq_value
        greater_values = [i for i in freq_array if i >= freq_value]
        # lower boundary freq value
        min_bound_val = max(less_values)
        # upper boundary freq value
        max_bound_val = min(greater_values)
        # lower boundary index
        min_bound_idx = list(freq_array).index(min_bound_val)
        max_bound_idx = list(freq_array).index(max_bound_val)
        line_prop = ProcessROI.determine_line_equation(point_index_1=min_bound_val,
                                                       point_index_2=max_bound_val,
                                                       point_val_1=self.new_dict['values'][min_bound_idx],
                                                       point_val_2=self.new_dict['values'][max_bound_idx])
        if min_bound_val == max_bound_val:
            current_nps = self.new_dict['values'][min_bound_idx]
        else:
            current_nps = line_prop[0] * freq_value + line_prop[1]
        # print('min bound value: ', self.new_dict['values'][min_bound_idx],
        #       '  freq_value:  ', self.new_dict['values'][max_bound_idx],
        #       'min bound index: ', min_bound_idx, '\n',
        #       'max bound value: ', max_bound_val, 'max bound index:  ', max_bound_idx)
        if current_nps < 0:
            current_nps = 0
        return current_nps

    def prepare_f_2(self, xy, a, b, c, d, e, f, g, h, k):
        """Auxiliar function for 2d-fitting"""
        i = xy // self.image_width_1  # reconstruct y coordinates
        j = xy % self.image_width_1  # reconstruct x coordinates
        out = i * a + j * b + i * j * c + i * j ** 2 * d + \
            i ** 2 * j * e + i ** 2 * j ** 2 * f + \
            i ** 2 * g + j ** 2 * h + k
        return out

    def create_pol_fit(self, array):
        
        """
        Create 2d fit of current ROI, either of first or of second order.
        
        :param array: ndarray (2d)
            Pixel array of current ROI.
        :return: ndarray (2d)
            Fitted pixel array.
        """
        
        self.pol_fit = []
        self.image_width_1 = array.shape[1]
        self.image_height_1 = array.shape[0]
        x = np.linspace(0, self.image_width_1 - 1, num=self.image_width_1)
        y = np.linspace(0, self.image_height_1 - 1, num=self.image_height_1)
        z = array[0: self.image_height_1, 0: self.image_width_1]
        xy = np.arange(z.size)  
        if self.fit_order == 1:
            mvfs = np.ravel(z)
            res = opt.curve_fit(self.prepare_f_1, xy, np.ravel(z))
            z_est = self.prepare_f_1(xy, *res[0])
        else:
            mvfs = np.ravel(z)
            res = opt.curve_fit(self.prepare_f_2, xy, np.ravel(z))
            z_est = self.prepare_f_2(xy, *res[0])

        self.pol_fit = z_est.reshape(self.image_height_1, self.image_width_1)
        # self.fitting = np.array(self.polyfit2d(x, y, z))
        # for item_fit_ind in range(self.fitting.shape[1]):
        #     self.pol_fit_sub = self.polyval2d(x, y, self.fitting[:, item_fit_ind])
        #     self.pol_fit.append(self.pol_fit_sub)
        return self.pol_fit

    def truncate_nps_freq(self, *, dict):
        
        """
        Truncate low NPS values at higher frequencies.
        
        :param dict: dict
            See return value of method compute_nps.
            
        :return: dict
            Dict truncated NPS values and respective frequencies.
            Keys : 'values', 'frequencies'.
        """
        
        """Takes as parameter dict: raw NPS-dict.
        Truncate higher frequencies with small respective NPS-values"""
        # select element greater than 10**(-4)
        truncated_nps = []
        for i in dict['values']:
            if i > self.trunc_percentage / 100 * np.max(dict['values']):
                truncated_nps.append(i)
            else:
                break
        # difference of length between normal and truncated NPS-lists
        # = last index of freq-list
        tr_idx = len(truncated_nps)
        # truncate freq-list
        truncated_freqs = dict['frequencies'][: tr_idx]
        new_dict = {'values': truncated_nps,
                    'frequencies': truncated_freqs}
        return new_dict

    @staticmethod
    def determine_line_equation(*,
                                point_val_1,
                                point_index_1,
                                point_val_2,
                                point_index_2):
        
        """
        Find slope and bias (shift) of the line containing two passed points.
        
        :param point_val_1: float
            y value of first point.
        :param point_index_1: float
            x value of first point.
        :param point_val_2: float
            y value of second point.
        :param point_index_2: float
            x value of second point.
        :return: tuple of floats
            Slope and bias of the aforementioned line.
        """
        
        # determination of line equation
        # based on two points of the line
        if point_index_1 == point_index_2:
            slope = 1
            shift = 0
        else:
            slope = (point_val_2 - point_val_1) / (point_index_2 - point_index_1)
            shift = point_val_1 - slope * point_index_1

        return slope, shift

    @staticmethod
    def collect_all_max_peaks_nps(dict):
        
        """
        Find all peak values and their indices and respective frequencies in passed 1d-list.
        
        :param dict: dict
            Keys : 'values', 'frequencies', 'AUC', 'Integral_of_2d_NPS'.
            Values : resp.: 1d-NPS values of current ROI, respective frequencies,
                            area under 1d-NPS profile, integral of 2d-NPS.
        :return: dict 
            Dict with all peaks info.
            Keys : 'values', 'indices', frequencies'
            Values : resp. list of peak values,
                           list of respective indices,
                           list of respective frequencies.
        """

        val_arr = dict['values']
        freq_arr = dict['frequencies']
        # counter for nps array items
        counter = 0
        # initialize max value - first threshold level
        max_value = 0
        # list to store all peaks values
        max_peaks_array = []
        # list to store all peaks indices
        max_ind_array = []
        # list to store respective frequencies
        resp_freq_max = []
        # auxiliary boolean variable
        switcher = True
        # initialize max index
        max_index = 0
        for item in val_arr:
            # if the next item has larger value than the previous
            # and if its value is more than the threshold level
            if item >= max_value:
                max_value = item
                max_index = counter
                # ability to store max value in the list
                switcher = True
            # if the next item is less than the previous (peak condition)
            elif switcher:
                max_peaks_array.append(max_value)
                max_ind_array.append(max_index)
                resp_freq_max.append(freq_arr[max_index])
                # prevent storing not peaks values
                switcher = False
            else:
                max_value = item
            counter += 1
        return {'values': max_peaks_array,#[1:],
                'indices': max_ind_array,#[1:],
                'frequencies': resp_freq_max}#[1:]}

    def handle_peak_info(self, peak_dict, all_val_arr, all_freq_arr):

        """
        Extract peak information from peak_dict.

        :param peak_dict: dict
            (See return value of static method collect_all_max_peaks_nps)
        :param all_val_arr: list
            List of values from which peak_dict has been built.
        :param all_freq_arr: list
            List of respective frequencies.
        :return: dict
         Dict with peak information:
            Keys : 'mean_value' - peak NPS value,
                   'mean_freq' - peak NPS frequency,
                   'left_dev' - frequency distance between peak freq. and
                                freq. at which NPS value falls under 60% of max value
                                to left side from peak.
                   'right_dev' - frequency distance between peak freq. and
                                 freq. at which NPS value falls under 60% of max value
                                 to right side from peak.
        """

        """Extract following info from peak_dict:
        - mean_value (i.e. absolute max peak value);
        - mean_freq (i.e. absolute max peak freq);
        - left deviation (freq, at which NPS-value falls underneath
                          the 60% of max value to the left side)
        - right deviation (the same, but to the right side)"""
        peak_val_arr = peak_dict['values']
        freq_arr = peak_dict['frequencies']
        left_dev = 'undefined'
        right_dev = 'undefined'
        # indices = peak_dict['indices']
        # if there are peaks
        no_peaks = False
        if len(peak_val_arr) > 0:
            only_right_dev = False
            # define max value and whether there is only right deviation
            try:
                if max(peak_val_arr) < max(all_val_arr) * 0.1:
                    mean_distr = max(all_val_arr)
                    index_max = list(all_val_arr).index(mean_distr)
                    only_right_dev = True
                else:
                    mean_distr = max(peak_val_arr)
                    index_max = list(all_val_arr).index(mean_distr)
                    if all([i > 0.6*mean_distr for i in all_val_arr[:index_max]]):
                        only_right_dev = True
            except ValueError:
                print('peak dict: ', peak_dict)
                print('all values: ', all_val_arr)
                print('file: ', self.basename)
        else:
            mean_distr = max(all_val_arr)
            index_max = list(all_val_arr).index(mean_distr)
            only_right_dev = True
            no_peaks = True

        # collect nps information
        if not no_peaks:
            try:
                # index_max_peak = peak_val_arr.index(mean_distr)
                mean_freq = all_freq_arr[list(all_val_arr).index(mean_distr)]
            except ValueError:
                print('file:  ', self.basename)
                print('peak dict: ', peak_dict)
                print('mean:   ', mean_distr)
        else:
            # index_max_peak = 0
            mean_freq = self.start_freq

        # right deviation
        for item in all_val_arr[index_max:]:
            if item < 0.6*mean_distr:
                r_dev_value = item
                resp_idx = list(all_val_arr).index(item)
                r_dev_freq = all_freq_arr[resp_idx]
                right_dev = r_dev_freq - mean_freq
                break
        # left deviation
        if only_right_dev:
            l_dev_value = 'undefined'
            l_dev_freq = 'undefined'
            left_dev = 'undefined'
        else:
            for item in all_val_arr[:index_max]:
                if item > 0.6 * mean_distr:
                    l_dev_value = item
                    resp_idx = list(all_val_arr).index(item)
                    l_dev_freq = all_freq_arr[resp_idx]
                    left_dev = mean_freq - l_dev_freq
                    break
        peak_info_dict = {'mean_value': mean_distr,
                          'mean_freq': mean_freq,
                          'left_dev': left_dev,
                          'right_dev': right_dev
                          }
        return peak_info_dict


if __name__ == '__main__':
    init_dict = {
                 'equal_ROIs_for each_image': True,
                 'useCentralCropping': True,
                 'crop_percentage': 38,
                 'useTruncation': False,
                 'trunc_percentage': 1,
                 'multipleFiles': True,
                 'useFitting': False,
                 'fitting_order': 2,
                 'image_height_in_mm': 'undefined',
                 'image_width_in_mm': 'undefined',
                 'pixel_size_mm': 0.781,
                 'extensions': ['.tif', '.dcm', '.png', '.jpg'],
                 'start_freq_range': 0,
                 'end_freq_range': 20,
                 'step': 0.01,
                 'main_window_width': 512,
                 'main_window_heigth': 512,
                 'list_of_indices_raw': [(0x0040, 0x0275, 0x0032, 0x1060),
                                         (0x7005, 0x100a),
                                         (0x0905, 0x1030),
                                         ],
                 'main_window_width_md': 350,
                 'main_window_height_md': 400,
                 'left_upper_corner_x_md': 30,
                 'left_upper_corner_y_md': 30,
                 'destroy_main_window': True,
                 'first_data_set': False
                 }

    # create base array dictionary for each image
    obj_array_dict = StartClass(suffixes=init_dict['extensions'],
                                list_of_indices_raw=init_dict['list_of_indices_raw'])

    # get program execution time
    start_time = time.time()

    # main window for GUI
    root1 = tk.Tk()

    # initial configuration of main window
    root1.geometry('%dx%d+%d+%d' % (init_dict['main_window_width'],
                                    init_dict['main_window_heigth'],
                                    100, 200))
    # size of the window 350px*(350+120)px; 120px - height of demonstration labels and slider
    root1.configure(background="black")  # initial background

    obj_gui = GUI(master=root1,
                  main_window_height=init_dict['main_window_heigth'],
                  main_window_width=init_dict['main_window_width'],
                  obj_arrays=obj_array_dict,
                  equal_ROIs_for_each_image=init_dict['equal_ROIs_for each_image'],
                  crop_perc=init_dict['crop_percentage'],
                  fit_order=init_dict['fitting_order'],
                  useFitting=init_dict['useFitting'],
                  im_height_in_mm=init_dict['image_height_in_mm'],
                  im_width_in_mm=init_dict['image_width_in_mm'],
                  pixel_size_in_mm=init_dict['pixel_size_mm'],
                  extensions=init_dict['extensions'],
                  trunc_percentage=init_dict['trunc_percentage'],
                  useCentralCropping=init_dict['useCentralCropping'],
                  useTruncation=init_dict['useTruncation'],
                  start_freq_range=init_dict['start_freq_range'],
                  end_freq_range=init_dict['end_freq_range'],
                  step=init_dict['step'],
                  multipleFiles=init_dict['multipleFiles'])
    obj_gui.grid()
    # roi_processing_start_time = time.time()
    obj_process_roi = ProcessROI(obj_arr=obj_array_dict,
                                 obj_roi=obj_gui,
                                 fit_order=init_dict['fitting_order'],
                                 crop_perc=init_dict['crop_percentage'],
                                 useFitting=init_dict['useFitting'],
                                 im_height_in_mm=init_dict['image_height_in_mm'],
                                 im_width_in_mm=init_dict['image_width_in_mm'],
                                 pixel_size_in_mm=init_dict['pixel_size_mm'],
                                 extensions=init_dict['extensions'],
                                 trunc_percentage=init_dict['trunc_percentage'],
                                 useCentralCropping=init_dict['useCentralCropping'],
                                 useTruncation=init_dict['useTruncation'],
                                 start_freq_range=init_dict['start_freq_range'],
                                 end_freq_range=init_dict['end_freq_range'],
                                 step=init_dict['step'],
                                 multipleFiles=init_dict['multipleFiles'],
                                 first_data_set=init_dict['first_data_set']
                                 )
    # menu option to start calculation
    obj_gui.menu.add_cascade(label='Final', menu=obj_gui.file)

    root1.mainloop()

    # print program execution time
    print('Program execution time is %f minutes' % ((time.time() - start_time) / 60))
    print('ROI processing time was %f minutes' % ((time.time() - roi_processing_start_time) / 60))



